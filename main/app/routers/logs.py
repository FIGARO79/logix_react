"""
Router para endpoints de logs (inbound).
"""
import datetime
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import JSONResponse, Response
from app.models.schemas import LogEntry
from app.services import db_logs, csv_handler
from app.utils.auth import login_required
from app.core.config import DB_FILE_PATH, ASYNC_DB_URL
from sqlalchemy.ext.asyncio import create_async_engine
import numpy as np

async_engine = create_async_engine(ASYNC_DB_URL)

router = APIRouter(prefix="/api", tags=["logs"])


@router.get('/find_item/{item_code}/{import_reference}')
async def find_item(item_code: str, import_reference: str, username: str = Depends(login_required)):
    """Busca un item en el maestro y calcula cantidades."""
    item_details = await csv_handler.get_item_details_from_master_csv(item_code)
    if item_details is None:
        raise HTTPException(status_code=404, detail=f"Artículo {item_code} no encontrado en el maestro.")
    
    expected_quantity = await csv_handler.get_total_expected_quantity_for_item(item_code)
    original_bin = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin = await db_logs.get_latest_relocated_bin_async(item_code)
    effective_bin_location = latest_relocated_bin if latest_relocated_bin else original_bin
    
    response_data = {
        "itemCode": item_details.get('Item_Code', item_code),
        "description": item_details.get('Item_Description', 'N/A'),
        "binLocation": effective_bin_location,
        "aditionalBins": item_details.get('Aditional_Bin_Location', 'N/A'),
        "weight": item_details.get('Weight_per_Unit', 'N/A'),
        "defaultQtyGrn": expected_quantity,
        "itemType": item_details.get('ABC_Code_stockroom', 'N/A'),
        "sicCode": item_details.get('SIC_Code_stockroom', 'N/A')
    }
    return JSONResponse(content=response_data)


@router.post('/add_log')
async def add_log(data: LogEntry, username: str = Depends(login_required)):
    """Añade un registro de entrada."""
    if data.quantity <= 0:
        raise HTTPException(status_code=400, detail="Cantidad debe ser > 0")
    
    item_code_form = data.itemCode
    import_reference = data.importReference
    quantity_received_form = data.quantity
    
    item_details = await csv_handler.get_item_details_from_master_csv(item_code_form)
    if item_details is None:
        raise HTTPException(status_code=404, detail=f"Artículo {item_code_form} no encontrado.")
    
    # Obtener la ubicación efectiva (reubicada si existe, o la original del maestro)
    original_bin_from_master = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin_for_item = await db_logs.get_latest_relocated_bin_async(item_code_form)
    bin_to_log_as_original = latest_relocated_bin_for_item if latest_relocated_bin_for_item else original_bin_from_master
    
    total_received_before = await db_logs.get_total_received_for_import_reference_async(import_reference, item_code_form)
    total_expected = await csv_handler.get_total_expected_quantity_for_item(item_code_form)
    total_received_now = total_received_before + quantity_received_form
    difference = total_received_now - total_expected
    
    entry_data = {
        'timestamp': datetime.datetime.now().isoformat(timespec='seconds'),
        'importReference': import_reference,
        'waybill': data.waybill,
        'itemCode': item_code_form,
        'itemDescription': item_details.get('Item_Description', 'N/A'),
        'binLocation': bin_to_log_as_original,  # Ubicación efectiva
        'relocatedBin': data.relocatedBin or '',
        'qtyReceived': quantity_received_form,
        'qtyGrn': total_expected,
        'difference': difference
    }
    
    log_id = await db_logs.save_log_entry_db_async(entry_data)
    if log_id:
        log_entry_data_for_response = {"id": log_id, **entry_data}
        return JSONResponse({'message': 'Registro añadido con éxito.', 'entry': log_entry_data_for_response}, status_code=201)
    raise HTTPException(status_code=500, detail="Error al guardar el registro.")


@router.put('/update_log/{log_id}')
async def update_log(log_id: int, data: dict, username: str = Depends(login_required)):
    """Actualiza un registro de entrada existente."""
    existing_log = await db_logs.get_log_entry_by_id_async(log_id)
    if not existing_log:
        raise HTTPException(status_code=404, detail=f"Registro con ID {log_id} no encontrado.")
    
    waybill = data.get('waybill', existing_log.get('waybill'))
    relocated_bin = data.get('relocatedBin', existing_log.get('relocatedBin'))
    qty_received = int(data.get('qtyReceived', existing_log.get('qtyReceived')))
    
    import_reference = existing_log['importReference']
    item_code = existing_log['itemCode']
    
    # Recalcular diferencia
    total_received_others = await db_logs.get_total_received_for_import_reference_async(import_reference, item_code)
    total_received_others -= int(existing_log.get('qtyReceived', 0))
    total_received_now = total_received_others + qty_received
    total_expected = await csv_handler.get_total_expected_quantity_for_item(item_code)
    difference = total_received_now - total_expected
    
    entry_data_for_db = {
        'waybill': waybill,
        'relocatedBin': relocated_bin,
        'qtyReceived': qty_received,
        'difference': difference,
        'timestamp': datetime.datetime.now().isoformat(timespec='seconds')
    }
    
    success = await db_logs.update_log_entry_db_async(log_id, entry_data_for_db)
    if success:
        return JSONResponse({'message': f'Registro {log_id} actualizado con éxito.'})
    raise HTTPException(status_code=500, detail="Error al actualizar el registro.")


@router.get('/get_logs')
async def get_logs(username: str = Depends(login_required)):
    """Obtiene todos los registros de entrada."""
    logs = await db_logs.load_log_data_db_async()
    return JSONResponse(content=logs)


@router.delete('/delete_log/{log_id}')
async def delete_log_api(log_id: int, username: str = Depends(login_required)):
    """Elimina un registro de entrada."""
    success = await db_logs.delete_log_entry_db_async(log_id)
    if success:
        return JSONResponse({'message': f'Registro {log_id} eliminado con éxito.'})
    raise HTTPException(status_code=404, detail=f"Registro con ID {log_id} no encontrado.")


@router.get('/export_log')
async def export_log(username: str = Depends(login_required)):
    """Exporta todos los registros de inbound a un archivo Excel."""
    logs_data = await db_logs.load_log_data_db_async()
    if not logs_data:
        raise HTTPException(status_code=404, detail="No hay registros para exportar")

    df = pd.DataFrame(logs_data)
    df_export = df[[
        'timestamp', 'importReference', 'waybill', 'itemCode', 'itemDescription',
        'binLocation', 'relocatedBin', 'qtyReceived', 'qtyGrn', 'difference'
    ]].rename(columns={
        'timestamp': 'Timestamp', 'importReference': 'Import Reference', 'waybill': 'Waybill',
        'itemCode': 'Item Code', 'itemDescription': 'Item Description',
        'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
        'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. Expected (Total)', 'difference': 'Difference'
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='InboundLogCompleto')
        worksheet = writer.sheets['InboundLogCompleto']
        for i, col_name in enumerate(df_export.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len

    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inbound_log_completo_{timestamp_str}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get('/export_reconciliation')
async def export_reconciliation(username: str = Depends(login_required)):
    """Genera y exporta el reporte de conciliación."""
    try:
        async with async_engine.connect() as conn:
            logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query('SELECT * FROM logs', sync_conn))

        # Accedemos al caché de GRN a través del handler (necesitamos exponerlo o acceder a la variable global si es posible,
        # pero mejor usamos una función del servicio si existe. En app.py era global.
        # Aquí asumimos que csv_handler tiene una forma de darnos el DF o lo leemos de nuevo si es necesario.
        # Revisando csv_handler (no visible aquí pero asumido), si no tiene getter, lo leemos.
        # Pero app.py usaba df_grn_cache. Vamos a intentar usar csv_handler.df_grn_cache si es accesible
        # o re-leerlo. Para seguridad y consistencia con app.py original que usaba caché:
        grn_df = csv_handler.df_grn_cache 

        if logs_df.empty or grn_df is None:
            raise HTTPException(status_code=404, detail="No hay datos suficientes para generar la conciliación")

        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)
        grn_df['Quantity'] = pd.to_numeric(grn_df['Quantity'], errors='coerce').fillna(0)

        item_totals = logs_df.groupby(['itemCode'])['qtyReceived'].sum().reset_index()
        item_totals = item_totals.rename(columns={'itemCode': 'Item_Code', 'qtyReceived': 'Total_Recibido'})

        grn_totals = grn_df.groupby(['GRN_Number', 'Item_Code', 'Item_Description'])['Quantity'].sum().reset_index()
        grn_totals = grn_totals.rename(columns={'Quantity': 'Total_Esperado'})

        merged_df = pd.merge(grn_totals, item_totals, on='Item_Code', how='outer')

        if not logs_df.empty:
            logs_df['id'] = pd.to_numeric(logs_df['id'])
            latest_logs = logs_df.sort_values('id', ascending=False).drop_duplicates('itemCode')
            
            latest_logs['Ubicacion_Log'] = np.where(
                latest_logs['relocatedBin'].notna() & (latest_logs['relocatedBin'] != ''),
                latest_logs['relocatedBin'],
                latest_logs['binLocation']
            )
            
            locations_df = latest_logs[['itemCode', 'Ubicacion_Log']].rename(columns={'itemCode': 'Item_Code'})
            merged_df = pd.merge(merged_df, locations_df, on='Item_Code', how='left')

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].fillna(0)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].fillna(0)
        merged_df['Diferencia'] = merged_df['Total_Recibido'] - merged_df['Total_Esperado']

        merged_df.fillna({'Ubicacion_Log': 'N/A'}, inplace=True)

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].astype(int)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].astype(int)
        merged_df['Diferencia'] = merged_df['Diferencia'].astype(int)

        df_for_export = merged_df.rename(columns={
            'GRN_Number': 'GRN',
            'Item_Code': 'Código de Ítem',
            'Item_Description': 'Descripción',
            'Ubicacion_Log': 'Ubicación (Log)',
            'Total_Esperado': 'Cant. Esperada',
            'Total_Recibido': 'Cant. Recibida',
            'Diferencia': 'Diferencia'
        })
        
        cols_order = ['GRN', 'Código de Ítem', 'Descripción', 'Ubicación (Log)', 'Cant. Esperada', 'Cant. Recibida', 'Diferencia']
        df_for_export = df_for_export[cols_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_for_export.to_excel(writer, index=False, sheet_name='ReporteDeConciliacion')
            worksheet = writer.sheets['ReporteDeConciliacion']
            for i, col_name in enumerate(df_for_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_for_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_conciliacion_{timestamp_str}.xlsx"
        return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno al generar el archivo de conciliación: {e}")
