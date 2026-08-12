"""
Router para endpoints de gestión de inventario y conteos administrativos.
"""

import datetime
from collections import defaultdict

from io import BytesIO
from urllib.parse import urlencode
from typing import Optional, Dict, Any, List
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from fastapi import APIRouter, Request, Depends, HTTPException, status
from fastapi.responses import RedirectResponse, Response
from app.core.responses import ORJSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, insert, update, text

from app.core.db import get_db
from app.services import csv_handler
from app.utils.auth import permission_required
from app.models.sql_models import (
    User,
    AppState,
    StockCount,
    CountSession,
    RecountList,
    SessionLocation,
    BinLocation,
    W2WInventorySnapshot,
    CycleCount,
)
from app.services.csv_to_db import sync_master_csv_to_db

import json
import os
from app.core.config import PROJECT_ROOT

# --- Inicialización ---
router = APIRouter(tags=["inventory"])


def get_slotting_aisles_data():
    """Carga los pasillos oficiales (aisles) y mapa de ubicación -> pasillo desde slotting_parameters.json."""
    json_path = os.path.join(PROJECT_ROOT, "static", "json", "slotting_parameters.json")
    if not os.path.exists(json_path):
        return [], {}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            storage = data.get("storage", {})
            bin_to_aisle = {}
            aisles_set = set()
            for bin_code, info in storage.items():
                aisle = info.get("aisle")
                if aisle:
                    clean_bin = str(bin_code).strip().upper()
                    clean_aisle = str(aisle).strip().upper()
                    bin_to_aisle[clean_bin] = clean_aisle
                    aisles_set.add(clean_aisle)
            return sorted(list(aisles_set)), bin_to_aisle
    except Exception as e:
        print(f"Error cargando slotting_parameters.json: {e}")
        return [], {}


class ZoneAssignmentPayload(BaseModel):
    user_id: int
    assigned_zones: str


@router.get("/api/admin/inventory/available_aisles")
async def get_available_aisles():
    """Retorna la lista oficial de pasillos (aisles) configurados en el mapa de slotting."""
    aisles, _ = get_slotting_aisles_data()
    return ORJSONResponse({"aisles": aisles})


@router.get("/api/admin/inventory/auditor_zones")
async def get_auditor_zones(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).order_by(User.username))
    users = result.scalars().all()
    aisles, _ = get_slotting_aisles_data()
    return ORJSONResponse([
        {
            "id": u.id,
            "username": u.username,
            "assigned_zones": u.assigned_zones or "",
            "is_approved": u.is_approved
        }
        for u in users
    ])


@router.post("/api/admin/inventory/assign_zones")
async def save_auditor_zones(payload: ZoneAssignmentPayload, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.id == payload.user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    user.assigned_zones = payload.assigned_zones.strip().upper()
    await db.commit()
    return ORJSONResponse({"message": f"Pasillos actualizados para {user.username}: {user.assigned_zones}"})


async def get_inventory_summary_stats(db: AsyncSession) -> Optional[Dict[str, Any]]:
    """Calcula y devuelve un resumen de estadísticas para el panel de admin de inventario."""
    summary: Dict[str, Any] = {
        "general": {
            "total_items_master": 0,
        },
        "stages": {},
    }

    try:
        # Asegurar caché actualizado
        await csv_handler.reload_cache_if_needed()

        # --- Estadísticas Generales (del maestro de items) ---
        if csv_handler.master_qty_map:
            total_items_with_stock = sum(
                1
                for qty in csv_handler.master_qty_map.values()
                if qty is not None and int(qty) > 0
            )  # type: ignore[arg-type]
            summary["general"]["total_items_master"] = total_items_with_stock

        # --- Estadísticas por Etapa ---
        for stage_num in range(1, 5):
            # Items contados en esta etapa
            stmt_items_counted = (
                select(func.count(func.distinct(StockCount.item_code)))
                .join(CountSession, StockCount.session_id == CountSession.id)
                .where(CountSession.inventory_stage == stage_num)
            )

            items_counted = (await db.execute(stmt_items_counted)).scalar() or 0

            # Si no se contó nada en esta etapa, podemos saltarla
            if items_counted == 0:
                continue

            # Total de unidades contadas
            stmt_total_units = (
                select(func.sum(StockCount.counted_qty))
                .join(CountSession, StockCount.session_id == CountSession.id)
                .where(CountSession.inventory_stage == stage_num)
            )

            total_units_counted = (await db.execute(stmt_total_units)).scalar() or 0

            # Calcular diferencias para esta etapa
            stmt_diff = (
                select(
                    StockCount.item_code,
                    func.sum(StockCount.counted_qty).label("total_counted"),
                )
                .join(CountSession, StockCount.session_id == CountSession.id)
                .where(CountSession.inventory_stage == stage_num)
                .group_by(StockCount.item_code)
            )

            counted_items_result = (await db.execute(stmt_diff)).all()
            counted_items_map = {
                row.item_code: row.total_counted for row in counted_items_result
            }

            items_with_discrepancy: int = 0
            for item_code, total_counted in counted_items_map.items():
                system_qty_raw = csv_handler.master_qty_map.get(item_code)
                system_qty: int = 0
                if system_qty_raw is not None:
                    try:
                        system_qty = int(float(system_qty_raw))
                    except (ValueError, TypeError):
                        system_qty = 0

                if total_counted != system_qty:
                    items_with_discrepancy += 1  # type: ignore[operator]

            # Precisión del conteo
            accuracy: float = 0.0
            if items_counted > 0:
                accuracy = (
                    (items_counted - items_with_discrepancy) / items_counted
                ) * 100  # type: ignore[operator]

            # Efectividad de Cobertura
            coverage_effectiveness: float = 0.0
            total_items_master_with_stock: int = summary["general"].get(
                "total_items_master", 0
            )  # type: ignore[union-attr]
            if total_items_master_with_stock > 0:
                items_correctly_counted: int = items_counted - items_with_discrepancy  # type: ignore[operator]
                coverage_effectiveness = (
                    items_correctly_counted / total_items_master_with_stock
                ) * 100

            # Guardar estadísticas de la etapa
            stage_stats: Dict[str, Any] = {
                "items_counted": items_counted,
                "total_units_counted": total_units_counted,
                "items_with_discrepancy": items_with_discrepancy,
                "accuracy": f"{accuracy:.2f}%",
                "coverage_effectiveness": f"{coverage_effectiveness:.2f}%",
            }
            summary["stages"][stage_num] = stage_stats  # type: ignore[index]

        # --- Items en lista de reconteo (para etapas futuras) ---
        stages_dict: Dict[int, Dict[str, Any]] = summary["stages"]
        for stage_to_check in range(2, 5):
            stmt_recount = select(func.count(RecountList.item_code)).where(
                RecountList.stage_to_count == stage_to_check
            )
            items_in_recount_list = (await db.execute(stmt_recount)).scalar() or 0

            if stage_to_check in stages_dict:
                stages_dict[stage_to_check]["items_in_recount_list"] = (
                    items_in_recount_list
                )
            elif items_in_recount_list > 0:
                # Si la etapa aún no tiene conteos pero ya hay lista de reconteo
                stages_dict[stage_to_check] = {
                    "items_in_recount_list": items_in_recount_list
                }

    except Exception as e:
        print(f"Error al calcular estadísticas de inventario: {e}")
        return None

    return summary


# ===== RUTAS DE ADMIN INVENTORY =====


@router.post("/admin/inventory/start_stage_1", name="start_inventory_stage_1")
async def start_inventory_stage_1(
    request: Request,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """Inicia un nuevo ciclo de inventario en Etapa 1."""

    try:
        print("Limpiando tablas de inventario para un nuevo ciclo...")
        await db.execute(update(CycleCount).values(count_id=None))
        await db.execute(delete(SessionLocation))
        await db.execute(delete(W2WInventorySnapshot))
        await db.execute(delete(StockCount))
        await db.execute(delete(CountSession))
        await db.execute(delete(RecountList))

        print("Tablas de inventario limpiadas.")

        # Sincronizar maestro de items desde CSV a DB
        print("Sincronizando maestro de items...")
        await sync_master_csv_to_db(db)
        await csv_handler.reload_cache_if_needed()
        print("Sincronización completada.")

        # Congelar snapshot estático de stock del sistema
        now_ts = datetime.datetime.now().isoformat(timespec="seconds")
        snapshot_records = []
        for code, sys_qty in csv_handler.master_qty_map.items():
            code_clean = str(code).upper().strip()
            desc = csv_handler.master_desc_map.get(code_clean, "N/A")
            bin_loc = csv_handler.master_bin_map.get(code_clean, "SYSTEM")
            cost = float(csv_handler.master_cost_map.get(code_clean, 0.0))
            snapshot_records.append(
                W2WInventorySnapshot(
                    session_id=0,
                    item_code=code_clean,
                    description=desc,
                    bin_location=bin_loc,
                    system_qty=float(sys_qty),
                    unit_cost=cost,
                    created_at=now_ts,
                )
            )
        if snapshot_records:
            db.add_all(snapshot_records)
            print(f"Snapshot estático de inventario congelado con {len(snapshot_records)} ítems.")

        # Actualizar estado
        stmt_update = (
            update(AppState)
            .where(AppState.key == "current_inventory_stage")
            .values(value="1")
        )
        await db.execute(stmt_update)

        await db.commit()

        query_params = urlencode(
            {
                "message": "Inventario reiniciado en Etapa 1. Todos los datos y contadores han sido reseteados."
            }
        )
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )
    except Exception as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )


async def get_app_setting(db: AsyncSession, key: str, default_value: str) -> str:
    """Obtiene una configuración de AppState o crea el valor por defecto si no existe."""
    stmt = select(AppState).where(AppState.key == key)
    res = await db.execute(stmt)
    setting = res.scalar_one_or_none()
    if not setting:
        setting = AppState(key=key, value=default_value)
        db.add(setting)
        await db.commit()
        return default_value
    return setting.value


async def process_stage_advance_logic(db: AsyncSession, next_stage: int) -> int:
    """Procesa el avance de etapa calculando diferencias con la toma más reciente por ítem (Rust)
    e introduciendo tolerancias de cantidad y valor monetario."""
    await csv_handler.reload_cache_if_needed()

    # 1. Obtenemos conteos físicos de etapas anteriores a next_stage
    stmt = (
        select(
            StockCount.item_code,
            StockCount.counted_location,
            StockCount.counted_qty,
            CountSession.inventory_stage,
        )
        .join(CountSession, StockCount.session_id == CountSession.id)
        .where(CountSession.inventory_stage < next_stage)
    )
    result = await db.execute(stmt)
    rows = result.fetchall()

    physical_counts = [
        (
            str(r.item_code).upper().strip(),
            str(r.counted_location or "N/A").upper().strip(),
            float(r.counted_qty or 0.0),
            int(r.inventory_stage),
        )
        for r in rows
    ]

    # 2. Stock en sistema desde maestro RAM
    system_stock = [
        (
            str(code).upper().strip(),
            "SYSTEM",
            float(qty),
            float(csv_handler.master_cost_map.get(code, 0.0)),
        )
        for code, qty in csv_handler.master_qty_map.items()
    ]

    # 3. Invocación de Rust para cálculo de diferencias respetando etapas
    try:
        import logix_rust_core
        diff_records = logix_rust_core.calculate_w2w_differences_rust(
            system_stock, physical_counts
        )
    except Exception as e:
        print(f"Error invocando Rust en avance de etapa: {e}. Usando fallback Python.")
        # Fallback en Python si Rust no estuviera compilado
        stage_map = {}
        counts_map = {}
        for item, loc, qty, stg in physical_counts:
            key = (item, loc)
            curr_stg = stage_map.get(key, -1)
            if stg > curr_stg:
                counts_map[key] = qty
                stage_map[key] = stg
            elif stg == curr_stg:
                counts_map[key] = counts_map.get(key, 0.0) + qty

        all_items = set(csv_handler.master_qty_map.keys()) | {item for item, loc in counts_map.keys()}
        diff_records = []
        for item in all_items:
            sys_q = float(csv_handler.master_qty_map.get(item, 0.0))
            cost = float(csv_handler.master_cost_map.get(item, 0.0))
            cnt_q = sum(q for (i, l), q in counts_map.items() if i == item)
            diff_q = cnt_q - sys_q
            diff_records.append({
                "item_code": item,
                "location": "SYSTEM",
                "system_qty": sys_q,
                "counted_qty": cnt_q,
                "diff_qty": diff_q,
                "unit_cost": cost,
                "diff_val": diff_q * cost,
                "status": "OK" if diff_q == 0 else ("SOBRANTE" if diff_q > 0 else "FALTANTE")
            })

    # 4. Agrupar por item_code y evaluar tolerancias en Rust para máxima velocidad
    qty_tol_str = await get_app_setting(db, "w2w_qty_tolerance", "0.02")
    val_tol_str = await get_app_setting(db, "w2w_val_tolerance", "10.00")
    qty_tolerance = float(qty_tol_str)
    val_tolerance = float(val_tol_str)

    items_for_recount = []
    try:
        import logix_rust_core
        raw_tuples = [
            (
                str(r["item_code"]),
                float(r["system_qty"]),
                float(r["counted_qty"]),
                float(r["diff_qty"]),
                float(r["unit_cost"]),
            )
            for r in diff_records
        ]
        items_for_recount = logix_rust_core.filter_recount_items_with_tolerances_rust(
            raw_tuples, qty_tolerance, val_tolerance, next_stage
        )
    except Exception as e:
        print(f"Error procesando tolerancias en Rust: {e}. Usando fallback Python.")
        item_aggregates = {}
        for rec in diff_records:
            item = rec["item_code"]
            sys_q = float(rec["system_qty"])
            cnt_q = float(rec["counted_qty"])
            diff_q = float(rec["diff_qty"])
            cost = float(rec["unit_cost"])
            
            if item not in item_aggregates:
                item_aggregates[item] = {
                    "system_qty": 0.0,
                    "counted_qty": 0.0,
                    "diff_qty": 0.0,
                    "unit_cost": cost
                }
            item_aggregates[item]["system_qty"] += sys_q
            item_aggregates[item]["counted_qty"] += cnt_q
            item_aggregates[item]["diff_qty"] += diff_q

        for code, agg in item_aggregates.items():
            diff_q = agg["diff_qty"]
            abs_diff_q = abs(diff_q)
            if abs_diff_q <= 0.0001:
                continue
                
            sys_q = agg["system_qty"]
            cost = agg["unit_cost"]
            diff_val = abs_diff_q * cost
            
            exceeds_qty = (abs_diff_q / sys_q) > qty_tolerance if sys_q > 0 else False
            exceeds_val = diff_val > val_tolerance
            
            if exceeds_qty or exceeds_val:
                items_for_recount.append(
                    {"item_code": code, "stage_to_count": next_stage, "status": "pending"}
                )

    # 5. Limpiar y recrear lista de reconteo para next_stage
    await db.execute(
        delete(RecountList).where(RecountList.stage_to_count == next_stage)
    )

    if items_for_recount:
        await db.execute(insert(RecountList), items_for_recount)

    return len(items_for_recount)


@router.post("/admin/inventory/advance/{next_stage}", name="advance_inventory_stage")
async def advance_inventory_stage(
    request: Request,
    next_stage: int,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """Avanza el inventario a la siguiente etapa (Redirect HTML)."""
    try:
        recount_count = await process_stage_advance_logic(db, next_stage)

        # Actualizar estado de la aplicación
        stmt_update = (
            update(AppState)
            .where(AppState.key == "current_inventory_stage")
            .values(value=str(next_stage))
        )
        await db.execute(stmt_update)
        await db.commit()

        message = f"Proceso completado. Etapa de inventario avanzada a {next_stage}. Se encontraron {recount_count} ítems con diferencias."
        query_params = urlencode({"message": message})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )

    except Exception as e:
        query_params = urlencode({"error": f"Error inesperado: {e}"})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )


@router.post("/admin/inventory/finalize", name="finalize_inventory")
async def finalize_inventory(
    request: Request,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """Finaliza el ciclo de inventario."""
    try:
        stmt_update = (
            update(AppState)
            .where(AppState.key == "current_inventory_stage")
            .values(value="0")
        )
        await db.execute(stmt_update)
        await db.commit()

        query_params = urlencode(
            {"message": "Ciclo de inventario finalizado y cerrado."}
        )
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )
    except Exception as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )


@router.get("/admin/inventory/report", name="generate_inventory_report")
async def generate_inventory_report(
    request: Request,
    db: AsyncSession = Depends(get_db),
    user: str = Depends(permission_required("inventory")),
):
    """Genera un reporte Excel completo y valorizado del inventario (con Rust + openpyxl)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        await csv_handler.reload_cache_if_needed()

        result = await db.execute(
            text("""
            SELECT sc.item_code, sc.item_description, sc.counted_location, cs.inventory_stage, sc.counted_qty
            FROM stock_counts sc
            JOIN count_sessions cs ON sc.session_id = cs.id
        """)
        )
        rows = result.fetchall()
        if not rows:
            query_params = urlencode(
                {"error": "No hay datos de conteo para generar un informe."}
            )
            return RedirectResponse(
                url=f"/admin/inventory?{query_params}",
                status_code=status.HTTP_302_FOUND,
            )

        # Mapa de tomas por item y etapa: item -> {stage: sum_qty}
        item_stage_counts: Dict[str, Dict[int, float]] = {}
        item_descriptions: Dict[str, str] = {}
        item_locations: Dict[str, set] = {}

        for r in rows:
            code = str(r.item_code).upper().strip()
            desc = str(r.item_description or "")
            loc = str(r.counted_location or "").upper().strip()
            stg = int(r.inventory_stage)
            qty = float(r.counted_qty or 0.0)

            item_descriptions[code] = desc
            if loc:
                item_locations.setdefault(code, set()).add(loc)

            if code not in item_stage_counts:
                item_stage_counts[code] = {}
            item_stage_counts[code][stg] = item_stage_counts[code].get(stg, 0.0) + qty

        # Todos los ítems involucrados
        all_item_codes = sorted(list(set(csv_handler.master_qty_map.keys()) | set(item_stage_counts.keys())))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Informe_Inventario_W2W"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "Ítem",
            "Descripción",
            "Ubicación",
            "Costo",
            "Sist",
            "Etapa 1",
            "Etapa 2",
            "Etapa 3",
            "Etapa 4",
            "Contado",
            "Diff",
            "Valor Diff",
        ]

        # Estilos Excel (Header Azul #1E4A74, Segoe UI)
        header_fill = PatternFill(start_color="1E4A74", end_color="1E4A74", fill_type="solid")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=9)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 2
        total_diff_val = 0.0
        for code in all_item_codes:
            desc = item_descriptions.get(code) or csv_handler.master_desc_map.get(code, "N/A")
            bin_sys = csv_handler.master_bin_map.get(code, "N/A")
            sys_qty = int(float(csv_handler.master_qty_map.get(code, 0.0)))
            cost = float(csv_handler.master_cost_map.get(code, 0.0))

            stg_map = item_stage_counts.get(code, {})
            c1 = int(stg_map.get(1)) if 1 in stg_map else None
            c2 = int(stg_map.get(2)) if 2 in stg_map else None
            c3 = int(stg_map.get(3)) if 3 in stg_map else None
            c4 = int(stg_map.get(4)) if 4 in stg_map else None

            # Determinación de Cantidad Final Contada (toma de etapa más alta existente)
            final_counted = sys_qty
            for stg in [4, 3, 2, 1]:
                if stg in stg_map:
                    final_counted = int(stg_map[stg])
                    break

            diff_qty = final_counted - sys_qty
            diff_val = diff_qty * cost

            row_data = [
                code,
                desc,
                bin_sys,
                cost,
                sys_qty,
                c1 if c1 is not None else "-",
                c2 if c2 is not None else "-",
                c3 if c3 is not None else "-",
                c4 if c4 is not None else "-",
                final_counted,
                diff_qty,
                diff_val,
            ]
            ws.append(row_data)

            # Alineaciones y formatos
            for c_idx in range(1, 13):
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.font = cell_font
                cell.border = thin_border

            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left")
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="left")
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=4).number_format = "$#,##0.00"
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=12).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=12).number_format = "$#,##0.00"

            row_idx += 1
            total_diff_val += diff_val
            row_idx += 1

        # Fila de Totales
        totals_row = [
            "TOTALES CONSOLIDADOS", "", "", "", "",
            "", "", "", "", "",
            "", total_diff_val
        ]
        ws.append(totals_row)
        tot_row_idx = row_idx
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=tot_row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10, bold=True)
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            cell.border = thin_border

        ws.cell(row=tot_row_idx, column=12).number_format = "$#,##0.00"

        # Ajuste dinámico de columnas
        for i, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            max_len = max(len(str(col_name)), 12)
            ws.column_dimensions[col_letter].width = float(max_len + 4)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"informe_final_inventario_w2w_{timestamp_str}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        print(f"Error generando el informe de inventario W2W: {e}")
        query_params = urlencode({"error": f"No se pudo generar el informe: {str(e)}"})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND
        )


@router.get("/api/export_recount_list/{stage_number}", name="export_recount_list")
async def export_recount_list(
    request: Request,
    stage_number: int,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """Exporta la lista de items a recontar para una etapa específica."""

    result = await db.execute(
        select(RecountList.item_code).where(RecountList.stage_to_count == stage_number)
    )
    items_to_recount = result.all()  # list of Row objects

    if not items_to_recount:
        if stage_number >= 4:
            return await generate_inventory_report(request, db, user)

        query_params = urlencode({"error": f"No hay ítems en la lista de reconteo para la Etapa {stage_number}."})
        return RedirectResponse(
            url=f"/admin/inventory?{query_params}",
            status_code=status.HTTP_302_FOUND,
        )

    # Importar la función para obtener detalles del item
    from app.services.csv_handler import get_item_details_from_master_csv

    import polars as pl
    import openpyxl

    enriched_data = []
    for row in items_to_recount:
        item_code = row.item_code
        details = await get_item_details_from_master_csv(item_code)
        if details:
            enriched_data.append(
                {
                    "Código de Item": item_code,
                    "Descripción": details.get("Item_Description", "N/A"),
                    "Ubicación en Sistema": details.get("Bin_1", "N/A"),
                }
            )
        else:
            enriched_data.append(
                {
                    "Código de Item": item_code,
                    "Descripción": "ITEM NO ENCONTRADO EN MAESTRO",
                    "Ubicación en Sistema": "N/A",
                }
            )

    df = pl.DataFrame(enriched_data)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reconteo_Etapa_{stage_number}"
    ws.append(df.columns)
    for row in df.iter_rows():
        ws.append(list(row))
    for i, col_name in enumerate(df.columns, start=1):
        col_data = df[col_name].cast(pl.Utf8, strict=False)
        max_len = max(col_data.str.len_chars().max() or 0, len(col_name)) + 2
        ws.column_dimensions[get_column_letter(i)].width = float(max_len)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lista_reconteo_etapa_{stage_number}_{timestamp_str}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ===== APIs PARA REACT ADMIN INVENTORY =====


@router.get("/api/admin/inventory/summary")
async def get_inventory_summary_api(
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Obtiene el resumen del estado del inventario."""
    stats = await get_inventory_summary_stats(db)

    # Obtener estado actual
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0

    # Obtener configuraciones de tolerancia
    qty_tolerance = await get_app_setting(db, "w2w_qty_tolerance", "0.02")
    val_tolerance = await get_app_setting(db, "w2w_val_tolerance", "10.00")

    return ORJSONResponse(
        content={
            "stage": current_stage,
            "stats": stats,
            "settings": {
                "w2w_qty_tolerance": float(qty_tolerance),
                "w2w_val_tolerance": float(val_tolerance),
            },
        }
    )


@router.post("/api/admin/inventory/start_stage_1")
async def start_inventory_stage_1_api(
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Inicia Etapa 1."""
    # Reset Current Stage to 1
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    if not stage_state:
        stage_state = AppState(key="current_inventory_stage", value="1")
        db.add(stage_state)
    else:
        stage_state.value = "1"

    # Limpiar tablas respetando FK constraints
    await db.execute(update(CycleCount).values(count_id=None))
    await db.execute(delete(SessionLocation))
    await db.execute(delete(W2WInventorySnapshot))
    await db.execute(delete(StockCount))
    await db.execute(delete(CountSession))
    await db.execute(delete(RecountList))

    await db.commit()
    return ORJSONResponse(
        content={"message": "Inventario Etapa 1 iniciado correctamente", "stage": 1}
    )


@router.post("/api/admin/inventory/advance_stage/{next_stage}")
async def advance_inventory_stage_api(
    next_stage: int,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Avanza etapa de inventario calculando diferencias con la toma más reciente por ítem (Rust)."""
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0

    if next_stage != current_stage + 1:
        raise HTTPException(
            status_code=400,
            detail=f"No se puede avanzar a la etapa {next_stage} desde la etapa {current_stage}",
        )

    recount_count = await process_stage_advance_logic(db, next_stage)

    stage_state.value = str(next_stage)
    await db.commit()
    return ORJSONResponse(
        content={
            "message": f"Inventario avanzado exitosamente a Etapa {next_stage}. {recount_count} ítems agregados a reconteo.",
            "stage": next_stage,
            "recount_items_count": recount_count,
        }
    )


@router.post("/api/admin/inventory/finalize")
async def finalize_inventory_api(
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Finaliza inventario."""
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    if stage_state:
        stage_state.value = "0"
        await db.commit()
    return ORJSONResponse(
        content={"message": "Inventario finalizado correctamente", "stage": 0}
    )


@router.get("/api/recount_list/active")
async def get_active_recount_list(
    db: AsyncSession = Depends(get_db),
    user: str = Depends(permission_required("inventory")),
):
    """API: Retorna la lista de ítems a recontar para la etapa activa actual."""
    await csv_handler.reload_cache_if_needed()

    # 1. Obtener etapa actual
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0

    if current_stage < 2:
        return ORJSONResponse(
            content={
                "stage": current_stage,
                "total": 0,
                "recounted_count": 0,
                "pending_count": 0,
                "items": [],
            }
        )

    # 2. Ítems asignados para reconteo en esta etapa (únicamente los pendientes)
    stmt_recount = select(RecountList.item_code).where(
        RecountList.stage_to_count == current_stage,
        RecountList.status == "pending"
    )
    recount_rows = (await db.execute(stmt_recount)).scalars().all()

    if not recount_rows:
        return ORJSONResponse(
            content={
                "stage": current_stage,
                "total": 0,
                "recounted_count": 0,
                "pending_count": 0,
                "items": [],
            }
        )

    # 3. Consultar qué ítems ya se han recontado en esta etapa
    stmt_recounted_in_stage = (
        select(
            StockCount.item_code,
            func.sum(StockCount.counted_qty).label("total_counted"),
        )
        .join(CountSession, StockCount.session_id == CountSession.id)
        .where(CountSession.inventory_stage == current_stage)
        .group_by(StockCount.item_code)
    )
    recounted_res = await db.execute(stmt_recounted_in_stage)
    recounted_map = {
        str(r.item_code).upper().strip(): float(r.total_counted or 0.0)
        for r in recounted_res.fetchall()
    }

    # 3.5 Consultar ubicaciones físicas registradas en Fase 1
    stmt_p1_locs = (
        select(
            StockCount.item_code,
            StockCount.counted_location,
        )
        .join(CountSession, StockCount.session_id == CountSession.id)
        .where(CountSession.inventory_stage == 1)
        .order_by(StockCount.id.desc())
    )
    p1_res = await db.execute(stmt_p1_locs)
    phase1_loc_map = {}
    for r in p1_res.fetchall():
        code_k = str(r.item_code).upper().strip()
        if code_k not in phase1_loc_map and r.counted_location:
            phase1_loc_map[code_k] = str(r.counted_location).upper().strip()

    # 4. Enriquecer lista con datos del maestro y ubicación de Fase 1
    items_list = []
    recounted_count = 0

    for code_raw in recount_rows:
        code = str(code_raw).upper().strip()
        details = await csv_handler.get_item_details_from_master_csv(code, db) or {}
        is_recounted = code in recounted_map
        if is_recounted:
            recounted_count += 1

        sys_loc = details.get("Bin_1", "N/A")
        p1_loc = phase1_loc_map.get(code)
        effective_loc = p1_loc if p1_loc else sys_loc

        items_list.append(
            {
                "item_code": code,
                "description": details.get("Item_Description", "N/A"),
                "bin_location": effective_loc,
                "system_location": sys_loc,
                "phase1_location": p1_loc or "N/A",
                "is_recounted": is_recounted,
                "counted_qty_in_stage": recounted_map.get(code, 0.0),
            }
        )

    # 3.5 Filtrar por pasillos (aisles) asignados al auditor si existen
    stmt_user = select(User).where(User.username == user)
    user_row = (await db.execute(stmt_user)).scalar_one_or_none()
    if user_row and user_row.assigned_zones:
        user_aisles = [z.strip().upper() for z in user_row.assigned_zones.split(",") if z.strip()]
        if user_aisles:
            _, bin_to_aisle = get_slotting_aisles_data()
            filtered_items = []
            for item in items_list:
                loc = str(item.get("bin_location", "")).upper().strip()
                item_aisle = bin_to_aisle.get(loc)
                match = (item_aisle in user_aisles) or any(loc.startswith(a) for a in user_aisles)
                if match:
                    filtered_items.append(item)
            items_list = filtered_items
            recounted_count = sum(1 for item in items_list if item["is_recounted"])

    items_list.sort(key=lambda x: (x["is_recounted"], x["item_code"]))

    return ORJSONResponse(
        content={
            "stage": current_stage,
            "total": len(items_list),
            "recounted_count": recounted_count,
            "pending_count": len(items_list) - recounted_count,
            "items": items_list,
        }
    )


from pydantic import BaseModel

class W2WCountPayload(BaseModel):
    session_id: int
    item_code: str
    counted_qty: int
    counted_location: str
    description: Optional[str] = "N/A"
    bin_location_system: Optional[str] = "N/A"
    timestamp: Optional[str] = None


@router.post("/api/w2w/save_count")
async def save_w2w_count_api(
    payload: W2WCountPayload,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API dedicada para guardar tomas de Inventario General W2W (sin alterar conteos cíclicos)."""
    # 1. Verificar que la sesión exista y esté activa
    stmt = select(CountSession).where(CountSession.id == payload.session_id)
    session_row = (await db.execute(stmt)).scalar_one_or_none()

    if not session_row or session_row.status != "in_progress":
        raise HTTPException(
            status_code=400,
            detail="La sesión de inventario W2W especificada no existe o ya está cerrada.",
        )

    # 1.5 Verificar que la ubicación exista en el maestro de slotting
    clean_location = payload.counted_location.upper().strip()
    stmt_bin = select(BinLocation).where(BinLocation.bin_code == clean_location)
    bin_row = (await db.execute(stmt_bin)).scalar_one_or_none()

    if not bin_row:
        raise HTTPException(
            status_code=400,
            detail=f"La ubicación '{clean_location}' no es una ubicación válida en el maestro de slotting.",
        )

    # 1.6 Verificar si la ubicación pertenece a los pasillos asignados al auditor
    stmt_user = select(User).where(User.username == user)
    user_row = (await db.execute(stmt_user)).scalar_one_or_none()
    if user_row and user_row.assigned_zones:
        allowed_aisles = [z.strip().upper() for z in user_row.assigned_zones.split(",") if z.strip()]
        if allowed_aisles:
            _, bin_to_aisle = get_slotting_aisles_data()
            loc_aisle = bin_to_aisle.get(clean_location)
            match = (loc_aisle in allowed_aisles) or any(clean_location.startswith(a) for a in allowed_aisles)
            if not match:
                raise HTTPException(
                    status_code=400,
                    detail=f"La ubicación '{clean_location}' (Pasillo {loc_aisle or 'N/A'}) no pertenece a sus pasillos autorizados ({', '.join(allowed_aisles)}).",
                )

    # 2. Normalizar código y obtener descripción si no se proveyó
    clean_code = payload.item_code.upper().strip()
    clean_desc = payload.description or "N/A"
    if clean_desc == "N/A" or not clean_desc:
        details = await csv_handler.get_item_details_from_master_csv(clean_code, db=db)
        if details:
            clean_desc = details.get("Item_Description", "N/A")
        else:
            clean_desc = "ITEM NO REGISTRADO EN MAESTRO"

    # 3. Guardar exclusivamente en StockCount (W2W)
    ts = payload.timestamp or datetime.datetime.now().isoformat(timespec="seconds")
    new_count = StockCount(
        session_id=payload.session_id,
        timestamp=ts,
        item_code=clean_code,
        item_description=clean_desc,
        counted_qty=payload.counted_qty,
        counted_location=payload.counted_location.upper().strip(),
        bin_location_system=payload.bin_location_system or "N/A",
        username=user,
    )
    db.add(new_count)
    await db.commit()
    await db.refresh(new_count)

    return ORJSONResponse(
        content={
            "message": "Conteo W2W registrado correctamente",
            "count_id": new_count.id,
            "session_id": payload.session_id,
            "item_code": clean_code,
            "inventory_stage": session_row.inventory_stage,
        }
    )


# ===== [NUEVO] ENDPOINTS DE WMS PROFESIONAL (TOLERANCIAS Y CONCILIACIÓN) =====


class W2WSettingsPayload(BaseModel):
    w2w_qty_tolerance: float
    w2w_val_tolerance: float


@router.get("/api/w2w/settings")
async def get_w2w_settings_api(
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Obtener configuraciones de tolerancia W2W."""
    qty_tolerance = await get_app_setting(db, "w2w_qty_tolerance", "0.02")
    val_tolerance = await get_app_setting(db, "w2w_val_tolerance", "10.00")
    return ORJSONResponse(
        content={
            "w2w_qty_tolerance": float(qty_tolerance),
            "w2w_val_tolerance": float(val_tolerance),
        }
    )


@router.post("/api/w2w/settings")
async def save_w2w_settings_api(
    payload: W2WSettingsPayload,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Guardar configuraciones de tolerancia W2W."""
    qty_str = str(payload.w2w_qty_tolerance)
    val_str = str(payload.w2w_val_tolerance)

    for key, val in [("w2w_qty_tolerance", qty_str), ("w2w_val_tolerance", val_str)]:
        stmt = select(AppState).where(AppState.key == key)
        res = await db.execute(stmt)
        setting = res.scalar_one_or_none()
        if not setting:
            db.add(AppState(key=key, value=val))
        else:
            setting.value = val

    await db.commit()
    return ORJSONResponse(content={"message": "Configuraciones W2W actualizadas correctamente"})


@router.post("/api/admin/inventory/settings")
async def update_inventory_settings_api(
    payload: Dict[str, str],
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Actualiza las configuraciones de tolerancia de inventario general en AppState."""
    for key, value in payload.items():
        if key in ["w2w_qty_tolerance", "w2w_val_tolerance"]:
            # Validar formato float
            try:
                float(value)
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Valor inválido para {key}")

            stmt = select(AppState).where(AppState.key == key)
            res = await db.execute(stmt)
            setting = res.scalar_one_or_none()
            if not setting:
                setting = AppState(key=key, value=value)
                db.add(setting)
            else:
                setting.value = value
    await db.commit()
    return ORJSONResponse(content={"message": "Configuraciones actualizadas correctamente"})


class ApprovePayload(BaseModel):
    item_code: str


@router.post("/api/admin/inventory/approve_item")
async def approve_inventory_item_api(
    payload: ApprovePayload,
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """API: Aprueba manualmente una discrepancia de inventario para que no requiera reconteo en la etapa actual."""
    # 1. Obtener la etapa activa
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0

    clean_code = payload.item_code.upper().strip()

    # 2. Buscar si está en la lista de reconteo de la etapa actual
    stmt = select(RecountList).where(
        RecountList.item_code == clean_code,
        RecountList.stage_to_count == current_stage
    )
    res = await db.execute(stmt)
    recount_item = res.scalar_one_or_none()

    if not recount_item:
        # Si no existe en la etapa actual, lo creamos directamente marcado como aprobado
        recount_item = RecountList(
            item_code=clean_code,
            stage_to_count=current_stage,
            status="manually_approved"
        )
        db.add(recount_item)
    else:
        recount_item.status = "manually_approved"

    await db.commit()
    return ORJSONResponse(
        content={
            "message": f"Código {clean_code} aprobado manualmente para la etapa {current_stage}."
        }
    )


@router.get("/api/admin/inventory/reconciliation")
async def get_inventory_reconciliation_api(
    user: str = Depends(permission_required("inventory")),
    db: AsyncSession = Depends(get_db),
):
    """
    API: Retorna la lista consolidada de discrepancias de inventario general
    con el historial de conteo por etapa, diferencias, costos y estado de aprobación.
    """
    await csv_handler.reload_cache_if_needed()

    # 1. Obtener etapa actual
    result = await db.execute(
        select(AppState).where(AppState.key == "current_inventory_stage")
    )
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0

    # 2. Cargar todas las tomas físicas de StockCount
    stmt_counts = (
        select(
            StockCount.item_code,
            StockCount.counted_qty,
            CountSession.inventory_stage
        )
        .join(CountSession, StockCount.session_id == CountSession.id)
    )
    res_counts = await db.execute(stmt_counts)
    counts_rows = res_counts.fetchall()

    # Agrupar conteos físicos por item_code y stage
    counts_by_item = defaultdict(lambda: defaultdict(float))
    for r in counts_rows:
        code = str(r.item_code).upper().strip()
        counts_by_item[code][int(r.inventory_stage)] += float(r.counted_qty or 0.0)

    # 3. Cargar RecountList para ver el estado de aprobaciones manuales/pendientes
    stmt_recount = select(RecountList).where(RecountList.stage_to_count == current_stage)
    res_recount = await db.execute(stmt_recount)
    recount_rows = res_recount.scalars().all()
    recount_map = {r.item_code.upper().strip(): r for r in recount_rows}

    # 4. Obtener tolerancias de AppState
    qty_tol_str = await get_app_setting(db, "w2w_qty_tolerance", "0.02")
    val_tol_str = await get_app_setting(db, "w2w_val_tolerance", "10.00")
    qty_tolerance = float(qty_tol_str)
    val_tolerance = float(val_tol_str)

    # 4.5 Cargar Snapshot Estático del Sistema si está disponible
    snaps = []
    try:
        res_snap = await db.execute(select(W2WInventorySnapshot))
        snaps = res_snap.scalars().all()
    except Exception as e:
        print(f"Aviso consultando snapshot W2W: {e}")

    if snaps:
        master_items_tuples = [
            (
                s.item_code,
                s.description or "ITEM NO CATALOGADO",
                s.bin_location or "N/A",
                float(s.system_qty),
                float(s.unit_cost),
            )
            for s in snaps
        ]
    else:
        master_items_tuples = [
            (
                code,
                csv_handler.master_desc_map.get(code, "ITEM NO CATALOGADO"),
                csv_handler.master_bin_map.get(code, "N/A"),
                float(csv_handler.master_qty_map.get(code, 0.0)),
                float(csv_handler.master_cost_map.get(code, 0.0)),
            )
            for code in csv_handler.master_qty_map.keys()
        ]

    # 5. Ejecutar consolidación y tolerancia acelerada en Rust
    try:
        import logix_rust_core
        count_records_tuples = [
            (str(r.item_code).upper().strip(), float(r.counted_qty or 0.0), int(r.inventory_stage))
            for r in counts_rows
        ]
        recount_statuses_map = {
            code: r.status for code, r in recount_map.items()
        }

        reconciliation_list = logix_rust_core.calculate_reconciliation_rust(
            master_items=master_items_tuples,
            count_records=count_records_tuples,
            recount_statuses=recount_statuses_map,
            _current_stage=current_stage,
            qty_tolerance=qty_tolerance,
            val_tolerance=val_tolerance,
        )
    except Exception as e:
        print(f"[RUST RECONCILIATION FALLBACK] Error executing Rust reconciliation: {e}")
        master_keys = set(csv_handler.master_qty_map.keys())
        counted_keys = set(counts_by_item.keys())
        all_keys = master_keys | counted_keys

        reconciliation_list = []
        for code_raw in all_keys:
            code = str(code_raw).upper().strip()
            if not code:
                continue

            sys_qty = float(csv_handler.master_qty_map.get(code, 0.0))
            cost = float(csv_handler.master_cost_map.get(code, 0.0))

            stage_counts = counts_by_item.get(code, {})
            c1 = stage_counts.get(1, None)
            c2 = stage_counts.get(2, None)
            c3 = stage_counts.get(3, None)
            c4 = stage_counts.get(4, None)

            final_counted = None
            for stg in [4, 3, 2, 1]:
                if stg in stage_counts:
                    final_counted = stage_counts[stg]
                    break
            
            is_counted = final_counted is not None
            final_counted_val = final_counted if is_counted else 0.0

            diff_qty = final_counted_val - sys_qty
            abs_diff_qty = abs(diff_qty)
            diff_val = diff_qty * cost
            abs_diff_val = abs(diff_val)

            status = "OK"
            if not is_counted:
                status = "NOT_COUNTED"
            elif abs_diff_qty > 0.0001:
                recount_item = recount_map.get(code)
                if recount_item:
                    if recount_item.status == "manually_approved":
                        status = "APPROVED_MANUAL"
                    else:
                        status = "PENDING_RECOUNT"
                else:
                    exceeds_qty = (abs_diff_qty / sys_qty) > qty_tolerance if sys_qty > 0 else False
                    exceeds_val = abs_diff_val > val_tolerance
                    if exceeds_qty or exceeds_val:
                        status = "PENDING"
                    else:
                        status = "APPROVED_AUTO"

            description = csv_handler.master_desc_map.get(code, "ITEM NO CATALOGADO")
            bin_location = csv_handler.master_bin_map.get(code, "N/A")

            reconciliation_list.append({
                "item_code": code,
                "description": description,
                "bin_location": bin_location,
                "system_qty": sys_qty,
                "cost": cost,
                "c1": c1,
                "c2": c2,
                "c3": c3,
                "c4": c4,
                "final_counted": final_counted_val,
                "diff_qty": diff_qty,
                "diff_val": diff_val,
                "status": status,
                "is_counted": is_counted
            })

        status_order = {
            "PENDING_RECOUNT": 0,
            "PENDING": 1,
            "APPROVED_MANUAL": 2,
            "APPROVED_AUTO": 3,
            "OK": 4
        }
        reconciliation_list.sort(
            key=lambda x: (status_order.get(x["status"], 5), -abs(x["diff_val"]))
        )

    return ORJSONResponse(
        content={
            "stage": current_stage,
            "items": reconciliation_list,
            "settings": {
                "w2w_qty_tolerance": qty_tolerance,
                "w2w_val_tolerance": val_tolerance,
            },
        }
    )
