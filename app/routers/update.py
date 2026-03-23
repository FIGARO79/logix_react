from fastapi import APIRouter, Request, Form, Depends, HTTPException, status, File, UploadFile, Response, BackgroundTasks
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse
import polars as pl
import os
import shutil
import json
import datetime
import numpy as np
from urllib.parse import urlencode
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete
from app.core.db import get_db
from app.models.sql_models import Log

# Importaciones relativas desde la estructura del proyecto
from app.core.config import (
    GRN_JSON_DATA_PATH, 
    PO_LOOKUP_JSON_PATH, 
    STOCK_QTY_CACHE_PATH,
    ITEM_MASTER_CSV_PATH,
    GRN_CSV_FILE_PATH,
    PICKING_CSV_PATH,
    RESERVATION_CSV_PATH,
    GRN_COLUMN_NAME_IN_CSV,
    ADMIN_PASSWORD
)
from app.services.csv_handler import load_csv_data
from app.utils.auth import login_required
from app.core.templates import templates

def np_encoder(obj):
    if isinstance(obj, np.generic):
        return obj.item()
    return str(obj)

router = APIRouter(
    prefix="",
    tags=["update"]
)

# --- Endpoint para la página de actualización (GET) ---
@router.get('/update', response_class=HTMLResponse)
async def update_files_get(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username  # Devuelve la redirección si el login falla
    
    return templates.TemplateResponse("update.html", {
        "request": request,
        "error": request.query_params.get('error'),
        "message": request.query_params.get('message')
    })

async def process_po_extractor_logic(file_path: str):
    """
    Procesa el archivo Excel de Purchase Order Extractor y genera el caché JSON.
    Esta función es compartida por la subida manual y el robot automático.
    """
    from app.core.config import PO_LOOKUP_JSON_PATH
    import datetime
    import json
    import numpy as np

    def np_encoder(obj):
        if isinstance(obj, np.generic):
            return obj.item()
        return str(obj)

    try:
        # Definir columnas base y opcionales
        base_cols = ["Waybill", "Import Ref Code", "Item Code", "Despatched Qty", "GRN Number"]
        opt_col = "Customer Reference"
        
        # Leer todo el Excel primero para verificar columnas
        df_full = pl.read_excel(file_path)
        
        # Filtrar columnas disponibles
        available_cols = [c for c in base_cols if c in df_full.columns]
        missing_base = [c for c in base_cols if c not in df_full.columns]
        
        if missing_base:
            return False, f"Faltan columnas obligatorias: {missing_base}"
            
        # Extraer datos base
        df_po = df_full.select(available_cols).select(pl.all().cast(pl.Utf8)).fill_null("")
        
        # Manejar columna opcional "Customer Reference"
        if opt_col in df_full.columns:
            df_po = df_po.with_columns(df_full.get_column(opt_col).cast(pl.Utf8).fill_null("").alias(opt_col))
        else:
            print(f"⚠️ Advertencia: Columna '{opt_col}' no encontrada en el Excel. Se usará vacía.")
            df_po = df_po.with_columns(pl.lit("").alias(opt_col))
        
        # LIMPIEZA CRÍTICA
        df_po = df_po.filter((pl.col("Waybill") != "") & (pl.col("Import Ref Code") != ""))
        
        # Normalizar datos
        df_po = df_po.with_columns([
            pl.col("Waybill").str.strip_chars().str.to_uppercase(),
            pl.col("Import Ref Code").str.strip_chars().str.to_uppercase(),
            pl.col("Item Code").str.strip_chars().str.to_uppercase(),
            pl.col("GRN Number").str.replace_all("/", ",").str.strip_chars(),
            pl.col(opt_col).str.strip_chars().str.to_uppercase()
        ])

        wb_lookup = {}
        ir_lookup = {}
        customer_ref_to_grn = {} # Mapeo: Customer Reference -> {grns, ir, waybill}

        # Procesar agrupado por Waybill
        for wb, group in df_po.group_by("Waybill"):
            wb_str = str(wb[0]) if isinstance(wb, tuple) else str(wb)
            first_row = group.row(0, named=True)
            items_list = []
            for row in group.iter_rows(named=True):
                items_list.append({
                    "item_code": row["Item Code"],
                    "qty": row["Despatched Qty"],
                    "grn": row["GRN Number"],
                    "customer_ref": row[opt_col]
                })
            
            wb_lookup[wb_str] = {
                "import_ref": first_row["Import Ref Code"],
                "items": items_list
            }

        # Generar mapeos basados en I.R. y Customer Reference
        for ir, group in df_po.group_by("Import Ref Code"):
            ir_str = str(ir[0]) if isinstance(ir, tuple) else str(ir)
            first_row = group.row(0, named=True)
            items_list = []
            for row in group.iter_rows(named=True):
                items_list.append({
                    "item_code": row["Item Code"],
                    "qty": row["Despatched Qty"],
                    "grn": row["GRN Number"],
                    "customer_ref": row[opt_col]
                })
                
                # Mapeo por Customer Reference (solo si existe)
                cust_ref = row[opt_col]
                if cust_ref:
                    if cust_ref not in customer_ref_to_grn:
                        customer_ref_to_grn[cust_ref] = {
                            "import_ref": ir_str,
                            "waybill": row["Waybill"],
                            "grns": set()
                        }
                    if row["GRN Number"]:
                        grns_in_row = set(g.strip().upper() for g in row["GRN Number"].split(',') if g.strip())
                        customer_ref_to_grn[cust_ref]["grns"].update(grns_in_row)
            
            ir_lookup[ir_str] = {
                "waybill": first_row["Waybill"],
                "items": items_list
            }
        
        # Convertir sets a listas para JSON
        for ref in customer_ref_to_grn:
            customer_ref_to_grn[ref]["grns"] = list(customer_ref_to_grn[ref]["grns"])

        lookup_data = {
            "wb_to_data": wb_lookup,
            "ir_to_data": ir_lookup,
            "customer_ref_to_data": customer_ref_to_grn,
            "updated_at": datetime.datetime.now().isoformat()
        }
        
        with open(PO_LOOKUP_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(lookup_data, f, indent=2, default=np_encoder)
        
        return True, "Caché de búsqueda generado correctamente."
    except Exception as e:
        print(f"Error procesando PO logic: {e}")
        return False, str(e)

from pydantic import BaseModel

class PORobotRequest(BaseModel):
    start_date: str
    end_date: str

# Variable global para el estado del robot en memoria
po_robot_status = {
    "status": "idle",
    "message": ""
}

@router.post('/api/run_po_robot', response_class=JSONResponse)
async def run_po_robot_api(
    payload: PORobotRequest,
    background_tasks: BackgroundTasks,
    username: str = Depends(login_required)
):
    """
    Dispara el robot de descarga de Purchase Order y luego procesa el archivo.
    """
    if not isinstance(username, str):
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"error": "Unauthorized"})

    async def execute_robot_task():
        global po_robot_status
        po_robot_status["status"] = "running"
        po_robot_status["message"] = f"Iniciando descarga para el periodo {payload.start_date} a {payload.end_date}..."
        
        from app.services.po_robot import run_po_robot
        from app.core.config import PO_EXTRACTOR_EXCEL_PATH
        
        # 1. Ejecutar descarga de forma asíncrona nativa
        success, msg = await run_po_robot(payload.start_date, payload.end_date)
        if not success:
            po_robot_status["status"] = "error"
            po_robot_status["message"] = f"Error en Robot: {msg}"
            print(f"❌ {po_robot_status['message']}")
            return

        # 2. Procesar el archivo
        success_proc, msg_proc = await process_po_extractor_logic(PO_EXTRACTOR_EXCEL_PATH)
        if success_proc:
            po_robot_status["status"] = "success"
            po_robot_status["message"] = f"Descarga y proceso completados con éxito. {msg_proc}"
            print(f"✅ Robot: {po_robot_status['message']}")
            # Recargar el caché de memoria general
            await load_csv_data()
        else:
            po_robot_status["status"] = "error"
            po_robot_status["message"] = f"Descarga OK pero error en proceso: {msg_proc}"
            print(f"❌ Robot: {po_robot_status['message']}")

    # Ejecutar en segundo plano para no bloquear al usuario
    background_tasks.add_task(execute_robot_task)
    
    return JSONResponse(content={"message": f"El robot ha sido activado para el periodo {payload.start_date} a {payload.end_date}. Consultando estado en tiempo real..."})

@router.get('/api/po_robot_status')
async def get_po_robot_status(username: str = Depends(login_required)):
    if not isinstance(username, str):
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"error": "Unauthorized"})
    # Log para depuración
    print(f"📡 [STATUS] Robot Status Check: {po_robot_status['status']} - {datetime.datetime.now().strftime('%H:%M:%S')}")
    return JSONResponse(content=po_robot_status)

# --- Endpoint para subir y procesar los archivos (POST) ---
@router.post('/api/update', response_class=JSONResponse)
async def update_files_post(
    request: Request,
    background_tasks: BackgroundTasks,
    item_master: UploadFile = File(None),
    grn_file: UploadFile = File(None),
    picking_file: UploadFile = File(None),
    reservation_file: UploadFile = File(None), # Nuevo campo para Reservas (Xdock)
    grn_excel: UploadFile = File(None),  # Nuevo campo para el Excel de Inbound
    po_extractor: UploadFile = File(None), # Nuevo campo para Purchase Order Extractor
    update_option_280: str = Form(None),
    selected_grns_280: str = Form(None),
    db: AsyncSession = Depends(get_db),
    username: str = Depends(login_required)
):
    if not isinstance(username, str):
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"error": "Unauthorized"})

    files_uploaded = False
    message = ""
    error = ""

    # Manejo del maestro de items
    if item_master and item_master.filename:
        with open(ITEM_MASTER_CSV_PATH, "wb") as buffer:
            shutil.copyfileobj(item_master.file, buffer)
        message += f'Archivo "{item_master.filename}" actualizado (Maestro). '
        files_uploaded = True

    # Manejo del archivo GRN (280)
    if grn_file and grn_file.filename:
        try:
            from app.services import reconciliation_service
            auto_snap = await reconciliation_service.auto_snapshot_before_update(db, username)
            if auto_snap:
                message += f"Snapshot de seguridad generado: {auto_snap}. "

            grn_bytes = grn_file.file.read()
            new_data_df = pl.read_csv(grn_bytes, infer_schema_length=0)
            
            if selected_grns_280:
                try:
                    selected_list = json.loads(selected_grns_280)
                    if selected_list:
                        new_data_df = new_data_df.filter(pl.col(GRN_COLUMN_NAME_IN_CSV).is_in(selected_list))
                except: pass

            if update_option_280 == 'combine' and os.path.exists(GRN_CSV_FILE_PATH):
                existing_data_df = pl.read_csv(GRN_CSV_FILE_PATH, infer_schema_length=0)
                new_grns = new_data_df.get_column(GRN_COLUMN_NAME_IN_CSV).unique()
                existing_data_df = existing_data_df.filter(~pl.col(GRN_COLUMN_NAME_IN_CSV).is_in(new_grns))
                combined_df = pl.concat([existing_data_df, new_data_df], how="vertical")
                combined_df.write_csv(GRN_CSV_FILE_PATH)
                message += f'Archivo "{grn_file.filename}" combinado. '
            else:
                new_data_df.write_csv(GRN_CSV_FILE_PATH)
                message += f'Archivo "{grn_file.filename}" reemplazado. '
            files_uploaded = True
        except Exception as e:
            error += f'Error procesando GRN: {str(e)}. '

    # Manejo del archivo de Reservas (AURRSLAMP0006)
    if reservation_file and reservation_file.filename:
        with open(RESERVATION_CSV_PATH, "wb") as buffer:
            shutil.copyfileobj(reservation_file.file, buffer)
        
        # [NUEVO] Generar caché rápido de Xdock en segundo plano
        from app.services.csv_handler import generate_reservation_cache
        background_tasks.add_task(generate_reservation_cache)
        
        message += f'Archivo "{reservation_file.filename}" actualizado (Xdock). '
        files_uploaded = True

    # Manejo del archivo de picking (240)
    if picking_file and picking_file.filename:
        with open(PICKING_CSV_PATH, "wb") as buffer:
            shutil.copyfileobj(picking_file.file, buffer)
        message += f'Archivo "{picking_file.filename}" actualizado (Picking). '
        files_uploaded = True

    # Manejo del archivo Excel de GRN (Inbound) -> Convertir a JSON
    if grn_excel and grn_excel.filename:
        try:
            excel_bytes = grn_excel.file.read()
            excel_df = pl.read_excel(excel_bytes)
            data_list = excel_df.to_dicts()
            with open(GRN_JSON_DATA_PATH, 'w', encoding='utf-8') as f:
                json.dump(data_list, f, indent=4, default=np_encoder)
            message += f'Archivo Excel "{grn_excel.filename}" procesado. '
            files_uploaded = True
            from app.services.grn_service import seed_grn_from_excel
            from app.core.db import AsyncSessionLocal
            async def run_sync():
                async with AsyncSessionLocal() as session:
                    await seed_grn_from_excel(session)
            background_tasks.add_task(run_sync)
        except Exception as e:
            error += f'Error Excel GRN: {str(e)}. '

    # Manejo del Purchase Order Extractor
    if po_extractor and po_extractor.filename:
        try:
            from app.core.config import PO_EXTRACTOR_EXCEL_PATH
            po_path = PO_EXTRACTOR_EXCEL_PATH
            with open(po_path, "wb") as buffer:
                shutil.copyfileobj(po_extractor.file, buffer)
            success, msg = await process_po_extractor_logic(po_path)
            if success:
                message += f"{msg} "
                files_uploaded = True
            else:
                error += f"Error PO Extractor: {msg}. "
        except Exception as e:
            error += f'Error PO Extractor (Crash): {str(e)}. '

    if files_uploaded:
        background_tasks.add_task(load_csv_data)
        message += " Procesamiento en segundo plano iniciado."

    if error:
        return JSONResponse(status_code=400, content={"error": error})
    return JSONResponse(content={"message": message or "No se subieron archivos."})


@router.post('/api/reload_cache', response_class=JSONResponse)
async def reload_cache_api(username: str = Depends(login_required)):
    """Fuerza la recarga de los datos CSV en la memoria RAM."""
    try:
        await load_csv_data()
        return {"message": "Caché de memoria RAM recargado correctamente"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al recargar caché: {e}")

# --- Endpoint para previsualizar las GRNs de un archivo ---
@router.post("/api/preview_grn_file")
async def preview_grn_file(file: UploadFile = File(...), username: str = Depends(login_required)):
    try:
        contents = await file.read()
        df = pl.read_csv(contents, infer_schema_length=0)
        if GRN_COLUMN_NAME_IN_CSV not in df.columns:
            return JSONResponse(status_code=400, content={"error": f"No se encontró la columna {GRN_COLUMN_NAME_IN_CSV}"})
        grns = sorted(df.get_column(GRN_COLUMN_NAME_IN_CSV).drop_nulls().unique().to_list())
        return JSONResponse(content={"grns": grns})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# --- Endpoint para la "Zona de Peligro" de limpiar la BD ---
@router.post('/api/clear_database')
async def clear_database_api(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    if password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "Contraseña incorrecta"})
    await db.execute(delete(Log))
    await db.commit()
    return JSONResponse(content={"message": "Base de datos de logs limpiada"})

@router.post('/clear_database')
async def clear_database(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    redirect_url = request.url_for('update_files_get')
    if password != ADMIN_PASSWORD:
        return RedirectResponse(url=f"{redirect_url}?error=Contraseña+incorrecta", status_code=302)
    await db.execute(delete(Log))
    await db.commit()
    return RedirectResponse(url=f"{redirect_url}?message=Base+de+datos+limpiada", status_code=302)

@router.post('/api/export_all_log')
async def export_all_log_api(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    if password != ADMIN_PASSWORD:
         return JSONResponse(status_code=401, content={"error": "Contraseña incorrecta"})
    try:
        from app.services import db_logs
        import polars as pl
        import openpyxl
        
        logs_data = await db_logs.load_all_logs_db_async(db)
        if not logs_data: return JSONResponse(status_code=404, content={"error": "No hay datos"})
        
        df = pl.DataFrame(logs_data)
        col_rename = {'timestamp': 'Date', 'importReference': 'Ref', 'itemCode': 'Item'}
        available = {k: v for k, v in col_rename.items() if k in df.columns}
        df_export = df.rename(available)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(df_export.columns)
        for row in df_export.iter_rows():
            ws.append(list(row))
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=backup_logs.xlsx"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})
