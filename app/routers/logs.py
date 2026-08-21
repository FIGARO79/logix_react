"""
Router para endpoints de logs (inbound).
"""

import datetime

from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from app.core.responses import ORJSONResponse
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.db import get_db
from app.models.schemas import LogEntry
from app.services import db_logs, csv_handler
from app.services.slotting_service import slotting_service
from app.utils.auth import login_required, permission_required
from app.core.config import ASYNC_DB_URL
from sqlalchemy.ext.asyncio import create_async_engine
from sqlalchemy import select

# Se mantiene el engine solo para pandas read_sql que requiere una conexión/engine
async_engine = create_async_engine(
    ASYNC_DB_URL,
    pool_pre_ping=True,
    pool_recycle=280,
)

router = APIRouter(prefix="/api", tags=["logs"])


from app.services.ai_slotting import ai_slotting


@router.get("/find_item/{item_code}/{import_reference}")
async def find_item(
    item_code: str,
    import_reference: str,
    username: str = Depends(permission_required(["stock", "inbound"])),
    db: AsyncSession = Depends(get_db),
):
    """Busca un item en el maestro y calcula cantidades con sugerencia IA."""
    # Uso de db=db para búsqueda rápida en SQL
    item_details = await csv_handler.get_item_details_from_master_csv(item_code, db=db)
    if item_details is None:
        raise HTTPException(
            status_code=404, detail=f"Artículo {item_code} no encontrado en el maestro."
        )

    # Obtener la cantidad esperada de la GRN (reporte 280) si está cargada y asociada a esta IR
    expected_quantity = (
        await csv_handler.get_expected_quantity_from_grn_for_import_ref(
            import_reference, item_code, db=db
        )
    ) or 0

    original_bin = item_details.get("Bin_1", "N/A")

    # Obtenemos la última reubicación para que aparezca como ubicación base si ya se movió
    latest_relocated_bin = await db_logs.get_latest_relocated_bin_async(db, item_code)

    # La ubicación base será la reubicada si existe, sino la del maestro
    effective_bin_location = (
        latest_relocated_bin if latest_relocated_bin else original_bin
    )

    # 1. Sugerencia de Slotting Dinámico (Algoritmo Tradicional)
    # Consolidación: si el item ya fue reubicado, usar esa ubicación como Bin_1
    # para que el algoritmo evalúe contra la posición real, no el maestro ERP.
    item_details_for_slotting = dict(item_details)
    if latest_relocated_bin:
        item_details_for_slotting["Bin_1"] = latest_relocated_bin

    traditional_suggested_bin = await slotting_service.get_suggested_bin(
        db, item_details_for_slotting
    )

    # 2. Sugerencia de IA (Aprendizaje Histórico)
    ai_predicted_bin = await ai_slotting.predict_best_bin(
        db=db,
        item_code=item_code,
        sic_code=item_details.get("SIC_Code_stockroom"),
        fallback_bin=traditional_suggested_bin,
    )

    # 3. VALIDACIÓN DE CAPACIDAD Y REGLAS DE ZONA PARA LA IA
    final_suggested_bin = ai_predicted_bin
    is_ai_prediction = ai_predicted_bin != traditional_suggested_bin

    if is_ai_prediction and ai_predicted_bin:
        config = await slotting_service._get_layout_config(db)
        storage = config.get("storage", {})
        zone_rules = config.get("zone_rules", {})
        mix_limits = config.get("mix_limits", {})

        pred_bin_info = storage.get(ai_predicted_bin.upper(), {})
        pred_zone = str(pred_bin_info.get("zone", "")).strip()
        try:
            pred_level = int(float(str(pred_bin_info.get("level", 0))))
        except (ValueError, TypeError):
            pred_level = 0

        minuteria_weight_max = float(zone_rules.get("minuteria_weight_max", 0.1))
        minuteria_zone = zone_rules.get("minuteria_zone", "Minuteria").strip()
        cantilever_kw = [
            k.strip().upper()
            for k in zone_rules.get("cantilever_keywords", "ROD, INTEGRAL STEEL").split(",")
            if k.strip()
        ]
        description = str(item_details.get("Item_Description", "")).upper()
        is_cantilever = any(kw in description for kw in cantilever_kw)

        weight = 0.0
        try:
            w_val = (
                item_details.get("Weight_per_Unit")
                or item_details.get("weight_per_unit")
                or item_details.get("weight", "0")
            )
            weight = float(str(w_val).replace(",", "")) if w_val else 0.0
        except Exception:
            pass

        # Regla 1: Ítem pesado o Cantilever NO puede ir a Minutería
        is_pred_minuteria = (
            pred_zone.upper() == "MINUTERIA"
            or pred_zone.upper() == minuteria_zone.upper()
        )
        if is_pred_minuteria and (weight > minuteria_weight_max or is_cantilever):
            final_suggested_bin = traditional_suggested_bin
            is_ai_prediction = False
        else:
            # Regla 2: Límite dinámico de capacidad por zona
            limit_minuteria = int(mix_limits.get("minuteria_max_skus", 3))
            limit_n2 = int(mix_limits.get("nivel2_max_skus", 6))
            limit_others = int(mix_limits.get("otros_niveles_max_skus", 4))

            if is_pred_minuteria:
                allowed_limit = limit_minuteria
            elif pred_level == 2:
                allowed_limit = limit_n2
            else:
                allowed_limit = limit_others

            occupancy = await slotting_service._get_bins_occupancy(db)
            current_skus = occupancy.get(ai_predicted_bin.upper(), 0)
            if current_skus >= allowed_limit:
                final_suggested_bin = traditional_suggested_bin
                is_ai_prediction = False

    # 4. Información de Cross-Docking (Xdock)
    xdock_data = await csv_handler.get_xdock_info(item_code, import_reference=import_reference)
    total_reserved = xdock_data.get("total", 0)
    raw_customers = xdock_data.get("customers", [])

    already_received = await db_logs.get_total_received_for_import_reference_async(
        db, import_reference, item_code
    )
    xdock_pending = max(0, total_reserved - already_received)

    # Deducir secuencialmente lo recibido para actualizar las cantidades pendientes por cliente
    xdock_customers = []
    rem_deduct = float(already_received)

    for c in raw_customers:
        if isinstance(c, dict):
            c_qty = float(c.get("qty", 0.0))
            if rem_deduct >= c_qty:
                rem_deduct -= c_qty
            else:
                pending_c_qty = c_qty - rem_deduct
                rem_deduct = 0.0
                c_copy = dict(c)
                c_copy["qty"] = pending_c_qty
                c_copy["original_qty"] = c_qty
                xdock_customers.append(c_copy)
        else:
            xdock_customers.append(c)

    # 5. ELIMINADO: Ya no sobreescribimos final_suggested_bin con "XDOCK" aquí.
    # El frontend manejará el indicador visual de XDOCK basado en xdock_pending.

    # 6. LIMPIEZA: No sugerir si el ítem ya está en la ubicación sugerida (comparando con el maestro)
    if final_suggested_bin == effective_bin_location:
        final_suggested_bin = None
        is_ai_prediction = False

    response_data = {
        "itemCode": item_details.get("Item_Code", item_code),
        "description": item_details.get("Item_Description", "N/A"),
        "binLocation": effective_bin_location,
        "suggestedBin": final_suggested_bin,
        "is_ai_prediction": is_ai_prediction,
        "xdockTotal": total_reserved,
        "xdockPending": xdock_pending,
        "xdockCustomers": xdock_customers,
        "aditionalBins": item_details.get("Aditional_Bin_Location", "N/A"),
        "physicalQty": str(item_details.get("Physical_Qty", "0")).replace(",", ""),
        "weight": item_details.get("Weight_per_Unit", "N/A"),
        "defaultQtyGrn": expected_quantity,
        "itemType": item_details.get("ABC_Code_stockroom", "N/A"),
        "sicCode": item_details.get("SIC_Code_stockroom", "N/A"),
        "frozenQty": item_details.get("Frozen_Qty", 0),
        "dateLastReceived": item_details.get("Date_Last_Received", "N/A"),
        "supersededBy": item_details.get("SupersededBy", "N/A"),
        "latestRelocatedBin": latest_relocated_bin,
        "expectedBreakdown": await csv_handler.get_expected_breakdown_by_item(
            item_code
        ),
    }
    return ORJSONResponse(content=response_data)


@router.post("/add_log")
async def add_log(
    data: LogEntry,
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Añade un registro de log (entrada de mercancía)."""
    item_code_form = data.itemCode.strip().upper()

    # Validar que el item existe usando SQL prioritariamente
    item_details = await csv_handler.get_item_details_from_master_csv(
        item_code_form, db=db
    )
    if not item_details:
        raise HTTPException(
            status_code=404, detail="El código de ítem no existe en el maestro."
        )

    expected_qty = data.qtyGrn
    if expected_qty is None:
        expected_qty = (
            await csv_handler.get_expected_quantity_from_grn_for_import_ref(
                data.importReference, item_code_form, db=db
            )
        ) or 0

    latest_relocated_bin = await db_logs.get_latest_relocated_bin_async(
        db, item_code_form
    )
    original_bin = item_details.get("Bin_1", "")
    effective_bin_location = (
        latest_relocated_bin if latest_relocated_bin else original_bin
    )

    prev_received = await db_logs.get_total_received_for_import_reference_async(
        db, data.importReference, item_code_form
    )
    cum_received = prev_received + data.quantity

    entry_data = data.dict()
    entry_data["username"] = username
    # Usar timestamp del frontend si viene, sino el del servidor (como fallback)
    if not entry_data.get("timestamp"):
        entry_data["timestamp"] = datetime.datetime.now().isoformat()
    entry_data["qtyGrn"] = expected_qty
    entry_data["qtyReceived"] = data.quantity
    entry_data["difference"] = cum_received - expected_qty

    entry_data["itemDescription"] = item_details.get("Item_Description", "")
    entry_data["binLocation"] = effective_bin_location

    if data.relocatedBin:
        await ai_slotting.learn_from_decision(
            db=db,
            item_code=item_code_form,
            final_bin=data.relocatedBin,
            sic_code=item_details.get("SIC_Code_stockroom"),
        )

    log_id = await db_logs.save_log_entry_db_async(db, entry_data)

    if log_id is not None and log_id > 0:
        return ORJSONResponse(
            content={"message": "Registro guardado correctamente", "id": log_id}
        )
    elif log_id == 0:
        raise HTTPException(
            status_code=409, detail="Registro duplicado detectado (client_id)."
        )
    else:
        raise HTTPException(
            status_code=500, detail="Error al guardar el registro en la base de datos."
        )


@router.get("/get_logs")
async def get_logs(
    version_date: Optional[str] = None,
    username: str = Depends(login_required),
    db: AsyncSession = Depends(get_db),
):
    """Obtiene los registros de log (activos por defecto o de una versión archivada)."""
    try:
        if version_date and version_date != "":
            logs = await db_logs.load_archived_log_data_db_async(db, version_date)
        else:
            logs = await db_logs.load_log_data_db_async(db)
        return ORJSONResponse(content=logs)
    except Exception as e:
        print(f"Error cargando logs: {e}")
        return ORJSONResponse(
            status_code=500, content={"error": "Error interno al cargar logs"}
        )


@router.delete("/delete_log/{log_id}")
async def delete_log(
    log_id: int,
    username: str = Depends(permission_required(["admin", "inbound"])),
    db: AsyncSession = Depends(get_db),
):
    """Elimina un registro de log."""
    success = await db_logs.delete_log_entry_db_async(db, log_id)
    if success:
        return ORJSONResponse(content={"message": "Registro eliminado"})
    else:
        raise HTTPException(status_code=404, detail="Registro no encontrado")


@router.put("/update_log/{log_id}")
async def update_log(
    log_id: int,
    data: dict,
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Actualiza un registro de log existente."""
    success = await db_logs.update_log_entry_db_async(db, log_id, data)
    if success:
        return ORJSONResponse(content={"message": "Registro actualizado correctamente"})
    else:
        raise HTTPException(
            status_code=404, detail="Registro no encontrado o error al actualizar"
        )


@router.post("/logs/archive")
async def archive_logs(
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Archiva los registros actuales."""
    archive_date = await db_logs.archive_current_logs_db_async(db)
    if archive_date:
        return ORJSONResponse(
            content={
                "message": "Registros archivados correctamente",
                "archive_date": archive_date,
            }
        )
    else:
        return ORJSONResponse(
            status_code=400,
            content={"message": "No hay registros activos para archivar"},
        )


from pydantic import BaseModel


class UnarchiveLogsRequest(BaseModel):
    version_date: str


@router.post("/logs/unarchive")
async def unarchive_logs(
    payload: UnarchiveLogsRequest,
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Desarchiva los registros de log correspondientes a una versión/fecha."""
    if not payload.version_date:
        raise HTTPException(status_code=400, detail="Debe especificar version_date")

    success = await db_logs.restore_archived_logs_db_async(db, payload.version_date)
    if success:
        return ORJSONResponse(
            content={"message": "Logs restaurados a activos correctamente"}
        )
    else:
        raise HTTPException(
            status_code=500, detail="Error interno al desarchivar los registros"
        )


@router.get("/logs/versions")
async def get_log_versions(
    username: str = Depends(login_required), db: AsyncSession = Depends(get_db)
):
    """Obtiene las fechas de las versiones archivadas."""
    versions = await db_logs.get_archived_versions_db_async(db)
    return ORJSONResponse(content=versions)


@router.get("/export_log")
async def export_log(
    timezone_offset: int = 0,
    version_date: Optional[str] = None,
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Exporta los registros de log a Excel con lógica de diferencia idéntica al frontend."""
    if version_date:
        logs = await db_logs.load_archived_log_data_db_async(db, version_date)
    else:
        logs = await db_logs.load_log_data_db_async(db)

    if not logs:
        raise HTTPException(status_code=404, detail="No hay registros para exportar")

    import polars as pl

    # Mapeo de columnas a español (coincidiendo con el frontend)
    col_map = {
        "id": "ID",
        "timestamp": "Fecha",
        "username": "Usuario",
        "importReference": "I.R.",
        "waybill": "Waybill",
        "itemCode": "Código Item",
        "itemDescription": "Descripción",
        "binLocation": "Ubicación",
        "relocatedBin": "Reubicación",
        "qtyReceived": "Cant. Recibida",
        "qtyGrn": "Cant. Esperada",
        "difference": "Diferencia",
    }

    cols_out = [
        "ID",
        "Fecha",
        "Usuario",
        "I.R.",
        "Waybill",
        "Código Item",
        "Descripción",
        "Ubicación",
        "Reubicación",
        "Cant. Recibida",
        "Cant. Esperada",
        "Diferencia",
    ]

    # ── APLICAR FORMATO DE FECHA (CON TIMEZONE OFFSET) ─────────────────────────
    enriched = []
    for log in logs:
        ts_raw = log.get("timestamp", "")
        formatted_date = ts_raw
        try:
            clean_ts = str(ts_raw).replace(" ", "T").replace("Z", "")
            dt_obj = datetime.datetime.fromisoformat(clean_ts)
            if dt_obj.tzinfo is None:
                dt_obj = dt_obj.replace(tzinfo=datetime.timezone.utc)
            local_dt = dt_obj - datetime.timedelta(minutes=timezone_offset)
            formatted_date = local_dt.replace(tzinfo=None)
        except Exception as e:
            print(f"Error procesando fecha en export: {e}")
            formatted_date = ts_raw

        log_entry = dict(log)
        log_entry["timestamp"] = formatted_date
        enriched.append(log_entry)

    # ─────────────────────────────────────────────────────────────────────────
    df_pl = pl.DataFrame(enriched, infer_schema_length=None)

    # Renombrar y seleccionar columnas
    available_cols = {k: v for k, v in col_map.items() if k in df_pl.columns}
    df_export = df_pl.rename(available_cols)

    # Seleccionar solo las columnas que existan en el mapeo y en el orden deseado
    final_cols = [c for c in cols_out if c in df_export.columns]
    df_export = df_export.select(final_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InboundLogs"

    # Cabeceras
    ws.append(df_export.columns)

    # Filas
    for row in df_export.iter_rows():
        ws.append(list(row))

    # Aplicar formato de fecha a la columna "Fecha"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime.datetime):
                cell.number_format = "DD/MM/YYYY HH:MM"

    # Auto-ajustar ancho de columnas
    for i, col_name in enumerate(df_export.columns, start=1):
        col_letter = get_column_letter(i)
        try:
            col_data = df_export[col_name].cast(pl.Utf8, strict=False)
            max_data = col_data.str.len_chars().max() or 0
        except:
            # Fallback para tipos Object (datetime)
            max_data = (
                max([len(str(v)) for v in df_export[col_name]])
                if len(df_export) > 0
                else 0
            )
        ws.column_dimensions[col_letter].width = float(
            max(int(max_data), len(col_name)) + 2
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = f"_{version_date}" if version_date else ""
    filename = (
        f"inbound_logs{suffix}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export_reconciliation")
async def export_reconciliation(
    timezone_offset: int = 0,
    archive_date: Optional[str] = None,
    snapshot_date: Optional[str] = None,
    username: str = Depends(permission_required("inbound")),
    db: AsyncSession = Depends(get_db),
):
    """Genera y exporta el reporte de conciliación (100% Polars, sin Pandas)."""
    import polars as pl
    from app.models.sql_models import ReconciliationHistory

    def _write_excel_polars(df: pl.DataFrame, sheet_name: str) -> bytes:
        """Convierte un DataFrame Polars a Excel con openpyxl, auto-ajustando anchos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Cabeceras
        ws.append(df.columns)
        # Filas
        for row in df.iter_rows():
            ws.append(list(row))

        # Aplicar formato de fecha a las celdas que contienen datetime
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime.datetime):
                    cell.number_format = "DD/MM/YYYY HH:MM"

        # Auto-ajustar ancho de columnas
        for i, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(i)
            try:
                col_data = df[col_name].cast(pl.Utf8, strict=False)
                max_data = col_data.str.len_chars().max() or 0
            except:
                # Fallback para tipos Object (datetime)
                max_data = (
                    max([len(str(v)) for v in df[col_name]]) if len(df) > 0 else 0
                )
            ws.column_dimensions[col_letter].width = float(
                max(int(max_data), len(col_name)) + 2
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    try:
        # ── RAMA SNAPSHOT ──────────────────────────────────────────────────────
        if snapshot_date:
            stmt = select(ReconciliationHistory).where(
                ReconciliationHistory.archive_date == snapshot_date
            )
            res = await db.execute(stmt)
            rows = res.scalars().all()

            if not rows:
                raise HTTPException(
                    status_code=404, detail="No se encontraron datos para este snapshot"
                )

            df_for_export = pl.DataFrame(
                [
                    {
                        "I.R.": r.import_reference,
                        "Waybill": r.waybill,
                        "GRN": r.grn,
                        "Código Item": r.item_code,
                        "Descripción": r.description,
                        "Ubicación": getattr(r, "bin_location", "") or "",
                        "Reubicado": getattr(r, "relocated_bin", "") or "",
                        "Cant. Esperada": int(r.qty_expected or 0),
                        "Cant. Recibida": int(r.qty_received or 0),
                        "Diferencia": int(r.difference or 0),
                        "Fecha": (
                            datetime.datetime.fromisoformat(
                                str(r.timestamp).replace("Z", "")
                            )
                            - datetime.timedelta(minutes=timezone_offset)
                        ).replace(tzinfo=None)
                        if r.timestamp
                        else None,
                    }
                    for r in rows
                ],
                infer_schema_length=None,
            )

            filename = f"snapshot_reconciliacion_{snapshot_date.replace(':', '-')}.xlsx"
            return Response(
                content=_write_excel_polars(df_for_export, "SnapshotConciliacion"),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

        # ── RAMA PRINCIPAL ─────────────────────────────────────────────────────
        else:
            # Usar la lógica centralizada del servicio
            from app.services import reconciliation_service

            result_data = await reconciliation_service.get_reconciliation_calculations(
                db, archive_date
            )

            if not result_data:
                raise HTTPException(
                    status_code=404, detail="No hay datos de conciliación para exportar"
                )

            final_df = pl.DataFrame(result_data, infer_schema_length=None)

            # Seleccionar y renombrar para el reporte Excel (manteniendo nombres de columnas del reporte)
            df_for_export = final_df.select(
                [
                    pl.col("Import_Reference").alias("I.R."),
                    pl.col("Waybill").alias("Waybill"),
                    pl.col("GRN").alias("GRN"),
                    pl.col("Codigo_Item").alias("Código Item"),
                    pl.col("Descripcion").alias("Descripción"),
                    pl.col("Ubicacion").alias("Ubicación"),
                    pl.col("Reubicado").alias("Reubicado"),
                    pl.col("Cant_Esperada").alias("Cant. Esperada"),
                    pl.col("Cant_Recibida").alias("Cant. Recibida"),
                    pl.col("Diferencia").alias("Diferencia"),
                    pl.col("Timestamp").alias("Fecha_ISO"),
                ]
            )

            # Ajustar la zona horaria en la columna Fecha usando Polars o mapeo manual
            def _adjust_tz(val):
                if not val:
                    return ""
                try:
                    clean_ts = str(val).replace(" ", "T").replace("Z", "")
                    dt = datetime.datetime.fromisoformat(clean_ts)
                    local_dt = dt - datetime.timedelta(minutes=timezone_offset)
                    return local_dt.replace(tzinfo=None)
                except:
                    return None

            df_for_export = df_for_export.with_columns(
                pl.col("Fecha_ISO")
                .map_elements(_adjust_tz, return_dtype=pl.Object)
                .alias("Fecha")
            ).drop("Fecha_ISO")

            utc_now = datetime.datetime.now(datetime.timezone.utc)
            client_time = utc_now - datetime.timedelta(minutes=timezone_offset)
            timestamp_str = client_time.strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_conciliacion_{timestamp_str}.xlsx"

            return Response(
                content=_write_excel_polars(df_for_export, "ReporteDeConciliacion"),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

    except HTTPException:
        raise
    except Exception as e:
        import traceback

        print(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"Error interno al generar el archivo de conciliación: {e}",
        )
