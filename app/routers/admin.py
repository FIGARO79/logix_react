"""
Router para endpoints administrativos simplificado y unificado.
"""
from fastapi import APIRouter, Request, Form, Depends, HTTPException, Body, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, ORJSONResponse, Response
from pydantic import BaseModel
from typing import List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, select, delete
from app.core.db import get_db
from app.utils.auth import (
    get_all_users, approve_user_by_id, delete_user_by_id, 
    reset_user_password, get_user_by_id, admin_login_required,
    permission_required
)
from app.models.sql_models import User, BinLocation, SlottingRule
from sqlalchemy import update
from app.core.config import ADMIN_PASSWORD, PROJECT_ROOT, SLOTTING_PARAMS_PATH
from app.core.templates import templates
from app.services.csv_handler import load_csv_data
from app.core.limiter import limiter
import orjson
import os

import datetime
from io import BytesIO

# Usaremos un solo router para evitar confusiones en main.py
router = APIRouter(prefix="/api/admin", tags=["admin"])

# --- Endpoints de Slotting ---

@router.get("/slotting-summary")
async def get_slotting_summary(admin: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Genera estadísticas reales cruzando el JSON con la DB."""
    try:
        # 1. Obtener layout desde JSON (caché rápido)
        config = {"storage": {}}
        if os.path.exists(SLOTTING_PARAMS_PATH):
            with open(SLOTTING_PARAMS_PATH, 'rb') as f:
                config = orjson.loads(f.read())
        
        storage = config.get("storage", {})
        
        # 2. Si el JSON está vacío, intentar obtener de SQL
        if not storage:
            res_bins = await db.execute(select(BinLocation))
            bins_sql = res_bins.scalars().all()
            storage = {b.bin_code: {"zone": b.zone, "aisle": b.aisle, "level": b.level, "spot": b.spot} for b in bins_sql}
            
        total_locations = len(storage)
        if total_locations == 0:
            return {"total": 0, "in_use": 0, "free": 0, "occupancy_pct": 0, "by_zone": {}}

        # Desglose por zona
        zones = {}
        for info in storage.values():
            z = info.get("zone", "Otras")
            zones[z] = zones.get(z, 0) + 1

        # Ocupación REAL: Bins con stock > 0 y total de ítems por bin
        query = text("SELECT bin_1, COUNT(*) as item_count FROM master_items WHERE physical_qty > 0 AND bin_1 IS NOT NULL GROUP BY bin_1")
        res = await db.execute(query)
        rows = res.all()
        bins_in_db = {str(row[0]).strip().upper(): row[1] for row in rows}

        # Cruzar con layout maestro (normalizar claves)
        storage_upper = {k.strip().upper(): v for k, v in storage.items()}
        matched_bins = {b: c for b, c in bins_in_db.items() if b in storage_upper}
        in_use_count = len(matched_bins)
        total_items_in_bins = sum(matched_bins.values())
        avg_items_per_bin = round(total_items_in_bins / in_use_count, 1) if in_use_count > 0 else 0

        # Saturación por zona y pasillo
        zone_items = {}
        aisle_items = {}
        for bin_code, item_count in matched_bins.items():
            info = storage_upper.get(bin_code, {})
            z = info.get("zone", "Otras")
            a = info.get("aisle", "?")
            zone_items[z] = zone_items.get(z, 0) + item_count
            if a and a != "nan":
                aisle_items[a] = aisle_items.get(a, 0) + item_count

        return {
            "total": total_locations,
            "in_use": in_use_count,
            "free": total_locations - in_use_count,
            "occupancy_pct": round((in_use_count / total_locations * 100), 1) if total_locations > 0 else 0,
            "by_zone": zones,
            "avg_items_per_bin": avg_items_per_bin,
            "total_items_in_bins": total_items_in_bins,
            "zones_by_items": dict(sorted(zone_items.items(), key=lambda x: x[1], reverse=True)),
            "top_aisles": dict(sorted(aisle_items.items(), key=lambda x: x[1], reverse=True)[:5])
        }
    except Exception as e:
        print(f"ERROR SUMMARY: {e}")
        return {"total": 0, "in_use": 0, "free": 0, "occupancy_pct": 0, "by_zone": {}}

@router.get("/slotting-config")
async def get_slotting_config(admin: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Retorna la configuración actual desde el JSON o reconstruye desde la DB si es necesario."""
    config = {"turnover": {}, "storage": {}}
    
    # 1. Intentar cargar desde JSON
    if os.path.exists(SLOTTING_PARAMS_PATH):
        try:
            with open(SLOTTING_PARAMS_PATH, 'rb') as f: 
                config = orjson.loads(f.read())
        except: pass

    # 2. Si el JSON está vacío o incompleto, reconstruir desde SQL
    if not config.get("storage") or not config.get("turnover"):
        print("⚡ [SLOTTING] Reconstruyendo configuración desde SQL...")
        
        res_bins = await db.execute(select(BinLocation))
        for b in res_bins.scalars().all():
            config["storage"][b.bin_code] = {
                "zone": b.zone, "aisle": b.aisle, "level": b.level, "spot": b.spot
            }
            
        res_rules = await db.execute(select(SlottingRule))
        for r in res_rules.scalars().all():
            config["turnover"][r.sic_code] = {
                "range": r.description or "", 
                "spot": r.ideal_spot
            }
            
        # Guardar en JSON para sincronizar cachés
        if config["storage"] or config["turnover"]:
            with open(SLOTTING_PARAMS_PATH, 'wb') as f: 
                f.write(orjson.dumps(config, option=orjson.OPT_INDENT_2))
    
    return config

@router.post("/slotting-config")
async def update_slotting_config(data: dict = Body(...), admin: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Guarda en JSON y sincroniza con la DB SQL."""
    # 1. Guardar JSON
    with open(SLOTTING_PARAMS_PATH, 'wb') as f: 
        f.write(orjson.dumps(data, option=orjson.OPT_INDENT_2))
    
    # 2. Sincronizar con SQL (Ubicaciones)
    storage = data.get("storage", {})
    if storage:
        # Enfoque conservador: Borrar y reinsertar es más simple para sincronización total
        # pero para rendimiento usaremos merges individuales si el set es pequeño.
        # Aquí asumimos actualización masiva desde panel.
        for code, info in storage.items():
            bin_code = code.strip().upper()
            stmt = select(BinLocation).where(BinLocation.bin_code == bin_code)
            res = await db.execute(stmt)
            existing = res.scalar_one_or_none()
            
            if existing:
                existing.zone = info.get("zone", "General")
                existing.aisle = info.get("aisle", "")
                existing.level = info.get("level", 0)
                existing.spot = info.get("spot", "Cold")
            else:
                db.add(BinLocation(
                    bin_code=bin_code,
                    zone=info.get("zone", "General"),
                    aisle=info.get("aisle", ""),
                    level=info.get("level", 0),
                    spot=info.get("spot", "Cold")
                ))

    # 3. Sincronizar con SQL (Reglas de Rotación)
    turnover = data.get("turnover", {})
    if turnover:
        for sic, info in turnover.items():
            sic_code = sic.strip().upper()
            stmt = select(SlottingRule).where(SlottingRule.sic_code == sic_code)
            res = await db.execute(stmt)
            existing = res.scalar_one_or_none()
            
            if existing:
                existing.ideal_spot = info.get("spot", "cold")
                existing.description = info.get("range", "")
            else:
                db.add(SlottingRule(
                    sic_code=sic_code,
                    ideal_spot=info.get("spot", "cold"),
                    description=info.get("range", "")
                ))

    await db.commit()
    return {"message": "Configuración guardada y sincronizada con base de datos"}

@router.get("/slotting-template")
async def get_slotting_template(admin: str = Depends(permission_required("inventory"))):
    data_list = []
    try:
        if os.path.exists(SLOTTING_PARAMS_PATH):
            with open(SLOTTING_PARAMS_PATH, 'rb') as f:
                storage = orjson.loads(f.read()).get('storage', {})
                for b, i in storage.items():
                    data_list.append({
                        "BIN": b, 
                        "ZONA": i.get('zone',''), 
                        "PASILLO": i.get('aisle',''), 
                        "NIVEL": i.get('level',0), 
                        "SPOT": i.get('spot','')
                    })
    except: pass
    
    import polars as pl
    import openpyxl
    df = pl.DataFrame(data_list if data_list else [{"BIN":"EJM-01-01", "ZONA":"ALMACEN", "PASILLO":"01", "NIVEL":1, "SPOT":"Hot"}])
    df = df.sort("BIN")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(df.columns)
    for row in df.iter_rows():
        ws.append(list(row))
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": "attachment; filename=layout_almacen.xlsx"})

@router.post("/slotting-upload")
async def upload_slotting_config(file: UploadFile = File(...), admin: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    try:
        import polars as pl
        file_bytes = await file.read()
        df = pl.read_excel(file_bytes)
        new_storage = {}
        
        # Limpiar tabla SQL antes de carga masiva
        await db.execute(delete(BinLocation))
        
        for r in df.iter_rows(named=True):
            b = str(r.get("BIN", "") or "").strip().upper()
            if b and b.lower() != "nan" and b.lower() != "none" and b != "":
                nivel = r.get("NIVEL")
                if nivel is None: nivel = 0
                
                info = {
                    "zone": str(r.get("ZONA", "") or "General"), 
                    "aisle": str(r.get("PASILLO", "") or ""), 
                    "level": nivel, 
                    "spot": str(r.get("SPOT", "") or "Cold")
                }
                new_storage[b] = info
                
                # Insertar en SQL
                db.add(BinLocation(
                    bin_code=b,
                    zone=info["zone"],
                    aisle=info["aisle"],
                    level=info["level"],
                    spot=info["spot"]
                ))
        
        # Sincronizar archivo JSON
        config = {"turnover": {}, "storage": new_storage}
        if os.path.exists(SLOTTING_PARAMS_PATH):
            with open(SLOTTING_PARAMS_PATH, 'rb') as f: 
                old_config = orjson.loads(f.read())
                config["turnover"] = old_config.get("turnover", {})
            
        with open(SLOTTING_PARAMS_PATH, 'wb') as f: 
            f.write(orjson.dumps(config, option=orjson.OPT_INDENT_2))
            
        await db.commit()
        return {"message": f"Cargadas {len(new_storage)} ubicaciones correctamente"}
    except Exception as e:
        await db.rollback()
        print(f"Error uploading slotting config: {e}")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

# --- Endpoints de Gestión de Usuarios ---

@router.get('/verify')
async def verify_admin_session(request: Request):
    """API: Verifica si hay una sesión activa de administrador."""
    if request.session.get("admin_logged_in"):
        return ORJSONResponse({"success": True})
    raise HTTPException(status_code=401, detail="No autorizado")

@router.get('/users')
async def get_admin_users_api(db: AsyncSession = Depends(get_db), admin: bool = Depends(admin_login_required)):
    users = await get_all_users(db)
    return ORJSONResponse(content=users)

@router.post('/login')
async def admin_login_api(request: Request, data: dict):
    if data.get('password') == ADMIN_PASSWORD:
        request.session['admin_logged_in'] = True
        return ORJSONResponse(content={"success": True})
    return ORJSONResponse(content={"success": False}, status_code=401)

@router.post('/approve/{user_id}')
async def approve_user_api(user_id: int, db: AsyncSession = Depends(get_db), admin: bool = Depends(admin_login_required)):
    success = await approve_user_by_id(db, user_id)
    if success:
        return ORJSONResponse(content={"success": True, "message": f"Usuario {user_id} aprobado"})
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

@router.post('/delete/{user_id}')
async def delete_user_api(user_id: int, db: AsyncSession = Depends(get_db), admin: bool = Depends(admin_login_required)):
    success = await delete_user_by_id(db, user_id)
    if success:
        return ORJSONResponse(content={"success": True, "message": f"Usuario {user_id} eliminado"})
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

@router.post('/reset_password/{user_id}')
async def admin_reset_password_api(user_id: int, new_password: str = Form(...), db: AsyncSession = Depends(get_db), admin: bool = Depends(admin_login_required)):
    success = await reset_user_password(db, user_id, new_password)
    if success:
        return ORJSONResponse(content={"success": True, "message": f"Contraseña restablecida para usuario {user_id}"})
    raise HTTPException(status_code=400, detail="Error al restablecer contraseña o contraseña no cumple requisitos")

@router.post('/permissions/{user_id}')
async def update_user_permissions_api(user_id: int, data: dict = Body(...), db: AsyncSession = Depends(get_db), admin: bool = Depends(admin_login_required)):
    perms_list = data.get('permissions', [])
    perms_str = ",".join(perms_list)
    
    stmt = update(User).where(User.id == user_id).values(permissions=perms_str)
    result = await db.execute(stmt)
    await db.commit()
    
    if result.rowcount > 0:
        return ORJSONResponse(content={"success": True, "message": "Permisos actualizados"})
    raise HTTPException(status_code=404, detail="Usuario no encontrado")

# Creamos un router vacío para compatibilidad con main.py si fuera necesario
api_router = APIRouter()
