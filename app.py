import asyncio
import os
import pandas as pd
from fastapi import FastAPI, Request, Form, Depends, HTTPException, status, File, UploadFile
from sqlalchemy.ext.asyncio import create_async_engine
from starlette.concurrency import run_in_threadpool
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.trustedhost import TrustedHostMiddleware
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import RedirectResponse, JSONResponse
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from werkzeug.security import generate_password_hash, check_password_hash
import datetime
import numpy as np
import sqlite3
import aiosqlite
import secrets
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional, List
from pydantic import BaseModel
from urllib.parse import urlencode
import shutil

# --- Cache para DataFrames ---
df_master_cache = None
df_grn_cache = None
# Map en memoria para cantidad por item (optimización)
master_qty_map = {}

# --- Configuración de Rutas ---
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))

DATABASE_FOLDER = os.path.join(PROJECT_ROOT, 'databases')
ITEM_MASTER_CSV_PATH = os.path.join(DATABASE_FOLDER, 'AURRSGLBD0250 - Item Stockroom Balance.csv')
GRN_CSV_FILE_PATH = os.path.join(DATABASE_FOLDER, 'AURRSGLBD0280 - Stock In Goods Inwards And Inspection.csv')
DB_FILE_PATH = os.path.join(PROJECT_ROOT, 'inbound_log.db')

# --- Configuración de la Base de Datos Asíncrona con SQLAlchemy ---
ASYNC_DB_URL = f"sqlite+aiosqlite:///{DB_FILE_PATH}"
async_engine = create_async_engine(ASYNC_DB_URL, echo=False)

# --- Configuración de Columnas ---
COLUMNS_TO_READ_MASTER = [
    'Item_Code', 'Item_Description', 'ABC_Code_stockroom', 'Physical_Qty','Frozen_Qty','Weight_per_Unit',
    'Bin_1', 'Aditional_Bin_Location','SupersededBy'
]
GRN_COLUMN_NAME_IN_CSV = 'GRN_Number'
COLUMNS_TO_READ_GRN = [GRN_COLUMN_NAME_IN_CSV, 'Item_Code', 'Quantity', 'Item_Description']

# --- Inicialización de FastAPI ---
app = FastAPI()


# Add this CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Permite solicitudes de cualquier origen.
    allow_credentials=True,
    allow_methods=["*"],  # Permite todos los métodos (POST, GET, etc).
    allow_headers=["*"],  # Permite todas las cabeceras.
)

# TrustedHostMiddleware procesa X-Forwarded-* cabeceras cuando la app está detrás
# de un proxy (por ejemplo PythonAnywhere). Debe añadirse antes de los middlewares
# que usan esas cabeceras para que request.url.scheme sea correcto.
app.add_middleware(TrustedHostMiddleware, allowed_hosts=["*"])

# --- Middleware CORREGIDO para forzar HTTPS y manejar 'scheme' ---
class SchemeMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        
        # 1. Determinar el 'scheme' correcto (http o https)
        # Por defecto, Uvicorn piensa que es http.
        scheme = request.scope.get('scheme', 'http')
        
        # Si la cabecera 'x-forwarded-proto' existe, esa es la verdad.
        if "x-forwarded-proto" in request.headers:
            scheme = request.headers['x-forwarded-proto']
        
        # 2. Forzar HTTPS *solo* en producción (PythonAnywhere)
        is_production = os.environ.get('PYTHONANYWHERE_DOMAIN')
        
        if is_production and scheme == 'http':
            # Si estamos en producción y la solicitud REAL es http, redirigimos a https
            https_url = str(request.url).replace("http://", "https://", 1)
            return RedirectResponse(https_url, status_code=status.HTTP_301_MOVED_PERMANENTLY)
        
        # 3. Si no redirigimos, nos aseguramos de que el 'scope' esté correcto para FastAPI
        request.scope['scheme'] = scheme
        
        # 4. Continuar con la solicitud
        response = await call_next(request)
        return response

app.add_middleware(SchemeMiddleware)
# Añadir HSTS en producción para que navegadores requieran HTTPS
class HSTSMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        is_production = os.environ.get('PYTHONANYWHERE_DOMAIN')
        scheme = request.scope.get('scheme', 'http')
        if is_production and scheme == 'https':
            # 2 años en segundos
            response.headers['Strict-Transport-Security'] = 'max-age=63072000; includeSubDomains; preload'
        return response

app.add_middleware(HSTSMiddleware)
# --- FIN DE LA CORRECCIÓN DE MIDDLEWARE ---

# --- Montar archivos estáticos ---
app.mount("/static", StaticFiles(directory=os.path.join(PROJECT_ROOT, "static")), name="static")

# --- Configuración de plantillas Jinja2 ---
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__)) # Asegúrate que esta línea está antes
templates = Jinja2Templates(directory=os.path.join(PROJECT_ROOT, "templates"))

# Helper para generar URLs seguras en producción (HTTPS)
def secure_url_for(request: Request, name: str, **path_params):
    """Genera URLs que usan HTTPS en producción y HTTP en desarrollo."""
    url = request.url_for(name, **path_params)
    # Si estamos en producción (PYTHONANYWHERE), fuerza HTTPS
    is_production = os.environ.get('PYTHONANYWHERE_DOMAIN')
    if is_production and str(url).startswith('http://'):
        url = str(url).replace('http://', 'https://', 1)
    return url

# Hacer el helper disponible en todas las plantillas Jinja2
templates.env.globals['secure_url_for'] = secure_url_for

# --- CONFIGURACIÓN DE SEGURIDAD ---
SECRET_KEY = 'una-clave-secreta-muy-dificil-de-adivinar'
UPDATE_PASSWORD = 'warehouse_admin_2025'

# --- Modelos Pydantic ---
class LogEntry(BaseModel):
    importReference: str
    waybill: str
    itemCode: str
    quantity: int
    relocatedBin: Optional[str] = ''

class Count(BaseModel):
    item_code: str
    quantity: int
    location: Optional[str] = 'N/A'

class StockCount(BaseModel):
    session_id: int
    item_code: str
    counted_qty: int
    counted_location: str
    description: Optional[str] = ''
    bin_location_system: Optional[str] = ''

class CloseLocationRequest(BaseModel):
    session_id: int
    location_code: str

class PickingAuditItem(BaseModel):
    code: str
    description: str
    qty_req: int
    qty_scan: int

class PickingAudit(BaseModel):
    order_number: str
    despatch_number: str
    customer_name: str
    status: str
    items: List[PickingAuditItem]

# --- LÓGICA DE LOGIN Y DECORADOR ---
def get_current_user(request: Request):
    return request.cookies.get("username")

def login_required(request: Request):
    if not get_current_user(request):
        return RedirectResponse(url='/login', status_code=status.HTTP_302_FOUND)
    return get_current_user(request)

# --- Funciones de Manejo de CSV ---
async def load_csv_data():
    global df_master_cache, df_grn_cache
    print("Cargando datos CSV en caché...")
    df_master_cache = await read_csv_safe(ITEM_MASTER_CSV_PATH, columns=COLUMNS_TO_READ_MASTER)
    df_grn_cache = await read_csv_safe(GRN_CSV_FILE_PATH, columns=COLUMNS_TO_READ_GRN)
    if df_master_cache is not None:
        print(f"Cargados {len(df_master_cache)} registros del maestro de items.")
        # Construir mapa en memoria item_code -> Physical_Qty (int o None)
        try:
            master_qty_map.clear()
            for _, row in df_master_cache.iterrows():
                code = row.get('Item_Code')
                raw_qty = row.get('Physical_Qty')
                qty_val = None
                if raw_qty not in (None, ''):
                    try:
                        qty_val = int(float(raw_qty))
                    except (ValueError, TypeError):
                        qty_val = None
                if code:
                    master_qty_map[code] = qty_val
        except Exception as e:
            print(f"Warning: no se pudo construir master_qty_map: {e}")
    if df_grn_cache is not None:
        print(f"Cargados {len(df_grn_cache)} registros del archivo GRN.")

async def read_csv_safe(file_path, columns=None):
    if not os.path.exists(file_path):
        print(f"Error CSV: Archivo no encontrado en {file_path}")
        return None
    try:
        df = await run_in_threadpool(pd.read_csv, file_path, usecols=columns, dtype=str, keep_default_na=True)
        df = df.replace({np.nan: None})
        return df
    except Exception as e:
        print(f"Error CSV: Error inesperado leyendo CSV {file_path}: {e}")
        return None

async def get_item_details_from_master_csv(item_code):
    if df_master_cache is None:
        raise HTTPException(status_code=500, detail="El maestro de items no está cargado.")
    result = df_master_cache[df_master_cache['Item_Code'] == item_code]
    return result.iloc[0].fillna('').to_dict() if not result.empty else None

async def get_total_expected_quantity_for_item(item_code_form):
    if df_grn_cache is None:
        return 0
    total_expected_quantity = 0
    result_df = df_grn_cache[df_grn_cache['Item_Code'] == item_code_form]
    if not result_df.empty:
        numeric_quantities = pd.to_numeric(result_df['Quantity'], errors='coerce').fillna(0)
        total_expected_quantity = int(numeric_quantities.sum())
    return total_expected_quantity

async def get_stock_data():
    if df_master_cache is not None:
        return df_master_cache[COLUMNS_TO_READ_MASTER]
    return None

# --- Funciones de Base de Datos SQLite ---
async def init_db():
    print("Inicializando y verificando el esquema de la base de datos...")
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            # --- Migración de la tabla de Logs ---
            conn.row_factory = aiosqlite.Row
            cursor = await conn.execute("PRAGMA table_info(logs);")
            columns = [row['name'] for row in await cursor.fetchall()]

            if 'packingListNumber' in columns and 'importReference' not in columns:
                print("Migrando esquema de la tabla 'logs': renombrando 'packingListNumber' a 'importReference'.")
                await conn.execute("ALTER TABLE logs RENAME COLUMN packingListNumber TO importReference;")

            # --- Tabla de Logs (existente) ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT NOT NULL, importReference TEXT NOT NULL DEFAULT '',
                    waybill TEXT, itemCode TEXT, itemDescription TEXT, binLocation TEXT,
                    relocatedBin TEXT, qtyReceived INTEGER, qtyGrn INTEGER, difference INTEGER
                )
            ''')

            # --- Tabla de Usuarios (existente) ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    is_approved INTEGER NOT NULL DEFAULT 0 -- 0 para pendiente, 1 para aprobado
                )
            ''')

            # --- Tablas para Sesiones de Conteo (Modificadas) ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS count_sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_username TEXT NOT NULL,
                    start_time TEXT NOT NULL,
                    end_time TEXT,
                    status TEXT NOT NULL DEFAULT 'in_progress', -- in_progress, completed
                    inventory_stage INTEGER NOT NULL DEFAULT 1 -- --- NUEVO: Se añade columna de etapa
                )
            ''')

            # --- NUEVO: Añadir columna 'inventory_stage' a 'count_sessions' si no existe (para BD antiguas) ---
            cursor = await conn.execute("PRAGMA table_info(count_sessions);")
            existing_cols_sessions = [row['name'] for row in await cursor.fetchall()]
            if 'inventory_stage' not in existing_cols_sessions:
                try:
                    await conn.execute("ALTER TABLE count_sessions ADD COLUMN inventory_stage INTEGER NOT NULL DEFAULT 1;")
                except aiosqlite.Error as e:
                    print(f"DB Warning: no se pudo añadir columna 'inventory_stage' a count_sessions: {e}")
            
            # --- NUEVO: Tabla 'app_state' para estado global del inventario ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS app_state (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            ''')
            
            # --- NUEVO: Inicializar el estado de la etapa de inventario a '1' si no está definida ---
            await conn.execute("INSERT OR IGNORE INTO app_state (key, value) VALUES ('current_inventory_stage', '1');")

            # --- NUEVO: Tabla 'recount_list' para la lista de tareas de reconteo ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS recount_list (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_code TEXT NOT NULL,
                    stage_to_count INTEGER NOT NULL DEFAULT 1,
                    status TEXT NOT NULL DEFAULT 'pending' -- pending, counted
                )
            ''')
            # --- FIN DE CAMBIOS DE ETAPA DE INVENTARIO ---


            await conn.execute('''
                CREATE TABLE IF NOT EXISTS session_locations ( 
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id INTEGER NOT NULL,
                    location_code TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'open', -- open, closed
                    closed_at TEXT,
                    FOREIGN KEY(session_id) REFERENCES count_sessions(id)
                )
            ''')
            
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS stock_counts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id INTEGER NOT NULL,
                    timestamp TEXT NOT NULL,
                    item_code TEXT NOT NULL,
                    item_description TEXT,
                    counted_qty INTEGER NOT NULL,
                    counted_location TEXT NOT NULL,
                    bin_location_system TEXT,
                    username TEXT,
                    FOREIGN KEY(session_id) REFERENCES count_sessions(id)
                )
            ''')

            # Asegurarse de que la columna 'username' exista para tablas creadas en versiones antiguas
            cursor = await conn.execute("PRAGMA table_info(stock_counts);")
            existing_cols = [row['name'] for row in await cursor.fetchall()]
            if 'username' not in existing_cols:
                try:
                    await conn.execute("ALTER TABLE stock_counts ADD COLUMN username TEXT;")
                except aiosqlite.Error as e:
                    print(f"DB Warning: no se pudo añadir columna 'username' a stock_counts: {e}")

            # --- Índices ---
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_importReference_itemCode ON logs (importReference, itemCode)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_session_id ON stock_counts (session_id)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_location_code ON session_locations (location_code)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_recount_item_stage ON recount_list (item_code, stage_to_count);") # --- NUEVO ÍNDICE ---

            # --- Nuevas Tablas para Auditoría de Picking ---
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS picking_audits (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    order_number TEXT NOT NULL,
                    despatch_number TEXT NOT NULL,
                    customer_name TEXT,
                    username TEXT NOT NULL,
                    timestamp TEXT NOT NULL,
                    status TEXT NOT NULL -- 'Completo' o 'Con Diferencia'
                )
            ''')

            await conn.execute('''
                CREATE TABLE IF NOT EXISTS picking_audit_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    audit_id INTEGER NOT NULL,
                    item_code TEXT NOT NULL,
                    description TEXT,
                    qty_req INTEGER NOT NULL,
                    qty_scan INTEGER NOT NULL,
                    difference INTEGER NOT NULL,
                    FOREIGN KEY(audit_id) REFERENCES picking_audits(id)
                )
            ''')
            
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_picking_audit_id ON picking_audit_items (audit_id)")

            # Tabla para tokens de restablecimiento de contraseña (un solo uso)
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS password_reset_tokens (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    token TEXT NOT NULL UNIQUE,
                    expires_at TEXT NOT NULL,
                    used INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(user_id) REFERENCES users(id)
                )
            ''')

            await conn.execute("CREATE INDEX IF NOT EXISTS idx_password_reset_token ON password_reset_tokens (token)")

            await conn.commit()
            print("Esquema de la base de datos verificado/actualizado con éxito.")
    except aiosqlite.Error as e:
        print(f"DB Error (init_db): {e}")

async def save_log_entry_db_async(entry_data):
    try:
        async with aiosqlite.connect(DB_FILE_PATH, timeout=10) as conn:
            sql = '''INSERT INTO logs (timestamp, importReference, waybill, itemCode, itemDescription,
                                     binLocation, relocatedBin, qtyReceived, qtyGrn, difference)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
            values = (
                entry_data.get('timestamp'), entry_data.get('importReference', ''), entry_data.get('waybill'),
                entry_data.get('itemCode'), entry_data.get('itemDescription'), entry_data.get('binLocation'),
                entry_data.get('relocatedBin'),
                entry_data.get('qtyReceived'), entry_data.get('qtyGrn'),
                entry_data.get('difference')
            )
            cursor = await conn.execute(sql, values)
            await conn.commit()
            return cursor.lastrowid
    except aiosqlite.Error as e:
        print(f"DB Error (save_log_entry_db_async): {e}")
        # No rollback needed with `async with` as it handles exceptions.
        return None

async def update_log_entry_db_async(log_id, entry_data_for_db):
    try:
        async with aiosqlite.connect(DB_FILE_PATH, timeout=10) as conn:
            sql = '''UPDATE logs SET
                        waybill = ?, relocatedBin = ?, qtyReceived = ?,
                        difference = ?, timestamp = ?
                     WHERE id = ?'''
            values = (
                entry_data_for_db.get('waybill'),
                entry_data_for_db.get('relocatedBin'),
                entry_data_for_db.get('qtyReceived'),
                entry_data_for_db.get('difference'),
                entry_data_for_db.get('timestamp'),
                log_id
            )
            cursor = await conn.execute(sql, values)
            await conn.commit()
            return cursor.rowcount > 0
    except aiosqlite.Error as e:
        print(f"DB Error (update_log_entry_db_async) para ID {log_id}: {e}")
        return False

async def load_log_data_db_async():
    logs = []
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            async with conn.cursor() as cursor:
                await cursor.execute("SELECT * FROM logs ORDER BY id DESC")
                rows = await cursor.fetchall()
                for row in rows:
                    logs.append(dict(row))
        return logs
    except aiosqlite.Error as e:
        print(f"DB Error (load_log_data_db_async): {e}")
        return []

async def get_log_entry_by_id_async(log_id):
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            async with conn.execute("SELECT * FROM logs WHERE id = ?", (log_id,)) as cursor:
                row = await cursor.fetchone()
                return dict(row) if row else None
    except aiosqlite.Error as e:
        print(f"DB Error (get_log_entry_by_id_async) para ID {log_id}: {e}")
        return None

async def get_total_received_for_import_reference_async(import_reference, item_code):
    total_received = 0
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            sql = "SELECT SUM(qtyReceived) FROM logs WHERE importReference = ? AND itemCode = ?"
            async with conn.execute(sql, (import_reference, item_code)) as cursor:
                result = await cursor.fetchone()
                if result and result[0] is not None: 
                    total_received = int(result[0])
    except aiosqlite.Error as e: 
        print(f"DB Error (get_total_received_for_import_reference_async): {e}")
    return total_received

async def delete_log_entry_db_async(log_id):
    try:
        async with aiosqlite.connect(DB_FILE_PATH, timeout=10) as conn:
            cursor = await conn.execute("DELETE FROM logs WHERE id = ?", (log_id,))
            await conn.commit()
            return cursor.rowcount > 0
    except aiosqlite.Error as e:
        print(f"DB Error (delete_log_entry_db_async) para ID {log_id}: {e}")
        return False

async def get_latest_relocated_bin_async(item_code):
    latest_bin = None
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            sql = '''
                SELECT relocatedBin FROM logs
                WHERE itemCode = ? AND relocatedBin IS NOT NULL AND relocatedBin != ''
                ORDER BY id DESC LIMIT 1
            '''
            async with conn.execute(sql, (item_code,)) as cursor:
                result = await cursor.fetchone()
                if result and result[0]:
                    latest_bin = result[0]
    except aiosqlite.Error as e:
        print(f"DB Error (get_latest_relocated_bin_async): {e}")
    return latest_bin

async def load_all_counts_db_async():
    counts = []
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            async with conn.cursor() as cursor:
                await cursor.execute("SELECT * FROM stock_counts ORDER BY id DESC")
                rows = await cursor.fetchall()
                for row in rows:
                    counts.append(dict(row))
        return counts
    except aiosqlite.Error as e:
        print(f"DB Error (load_all_counts_db_async): {e}")
        return []

# --- Endpoints de la API ---

# --- Endpoints de Sesiones de Conteo ---

@app.post("/api/sessions/start", status_code=status.HTTP_201_CREATED)
async def start_new_session(username: str = Depends(login_required)):
    """Inicia una nueva sesión de conteo para el usuario actual."""
    print(f"Attempting to start a new session for user: {username}")
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row # Añadir row_factory para leer la etapa
        try:
            # --- NUEVO: Obtener la etapa de inventario global actual ---
            cursor_stage = await conn.execute("SELECT value FROM app_state WHERE key = 'current_inventory_stage'")
            stage_row = await cursor_stage.fetchone()
            current_stage = int(stage_row['value']) if (stage_row and stage_row['value']) else 1 # Default a 1
            print(f"Global inventory stage is: {current_stage}")
            # --- FIN NUEVO ---

            # Opcional: Finalizar sesiones anteriores del mismo usuario
            print("Closing previous sessions...")
            await conn.execute(
                "UPDATE count_sessions SET status = 'completed', end_time = ? WHERE user_username = ? AND status = 'in_progress'",
                (datetime.datetime.now().isoformat(timespec='seconds'), username)
            )
            print("Previous sessions closed.")

            # Crear nueva sesión
            print(f"Creating new session for stage {current_stage}...") # Modificado
            cursor = await conn.execute(
                # --- MODIFICADO: Añadir inventory_stage ---
                "INSERT INTO count_sessions (user_username, start_time, status, inventory_stage) VALUES (?, ?, ?, ?)",
                (username, datetime.datetime.now().isoformat(timespec='seconds'), 'in_progress', current_stage)
            )
            await conn.commit()
            session_id = cursor.lastrowid
            print(f"New session created with ID: {session_id} for stage {current_stage}") # Modificado
            
            # --- MODIFICADO: Devolver la etapa al frontend ---
            return {"session_id": session_id, "inventory_stage": current_stage, "message": f"Sesión {session_id} (Etapa {current_stage}) iniciada."}
        
        except aiosqlite.Error as e:
            print(f"Database error in start_new_session: {e}")
            raise HTTPException(status_code=500, detail=f"Error de base de datos: {e}")

@app.get("/api/sessions/active")
async def get_active_session(username: str = Depends(login_required)):
    """Obtiene la sesión de conteo activa para el usuario."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute(
            "SELECT * FROM count_sessions WHERE user_username = ? AND status = 'in_progress' ORDER BY start_time DESC LIMIT 1",
            (username,)
        )
        session = await cursor.fetchone()
        if session:
            return dict(session)
        return JSONResponse(content={"message": "No hay sesión de conteo activa."}, status_code=404)

@app.post("/api/sessions/{session_id}/close")
async def close_session(session_id: int, username: str = Depends(login_required)):
    """Cierra una sesión de conteo."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        # Verificar que la sesión pertenece al usuario
        cursor = await conn.execute(
            "SELECT id FROM count_sessions WHERE id = ? AND user_username = ?", (session_id, username)
        )
        if not await cursor.fetchone():
            raise HTTPException(status_code=403, detail="No tienes permiso para cerrar esta sesión.")

        await conn.execute(
            "UPDATE count_sessions SET status = 'completed', end_time = ? WHERE id = ?",
            (datetime.datetime.now().isoformat(timespec='seconds'), session_id)
        )
        await conn.commit()
        return {"message": f"Sesión {session_id} cerrada con éxito."}

@app.post("/api/locations/close")
async def close_location(data: CloseLocationRequest, username: str = Depends(login_required)):
    """Marca una ubicación como 'cerrada' para una sesión de conteo."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        # Verificar que la sesión existe y pertenece al usuario
        cursor = await conn.execute(
            "SELECT id FROM count_sessions WHERE id = ? AND user_username = ? AND status = 'in_progress'",
            (data.session_id, username)
        )
        if not await cursor.fetchone():
            raise HTTPException(status_code=403, detail="La sesión no es válida o no te pertenece.")

        # Insertar o actualizar el estado de la ubicación
        await conn.execute(
            """
            INSERT INTO session_locations (session_id, location_code, status, closed_at)
            VALUES (?, ?, 'closed', ?)
            """,
            (data.session_id, data.location_code, datetime.datetime.now().isoformat(timespec='seconds'))
        )
        await conn.commit()
        return {"message": f"Ubicación {data.location_code} cerrada para la sesión {data.session_id}."}

@app.get("/api/sessions/{session_id}/locations")
async def get_session_locations(session_id: int, username: str = Depends(login_required)):
    """Obtiene el estado de todas las ubicaciones para una sesión."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        # Verificar que la sesión pertenece al usuario
        cursor = await conn.execute(
            "SELECT id FROM count_sessions WHERE id = ? AND user_username = ?", (session_id, username)
        )
        if not await cursor.fetchone():
            raise HTTPException(status_code=403, detail="No tienes permiso para ver esta sesión.")

        cursor = await conn.execute(
            "SELECT location_code, status FROM session_locations WHERE session_id = ?", (session_id,)
        )
        locations = await cursor.fetchall()
        return [dict(row) for row in locations]

@app.get("/api/sessions/{session_id}/counts/{location_code}")
async def get_counts_for_location(session_id: int, location_code: str, username: str = Depends(login_required)):
    """Obtiene todos los conteos para una ubicación específica en una sesión."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        # Verificar permiso
        cursor = await conn.execute(
            "SELECT id FROM count_sessions WHERE id = ? AND user_username = ?", (session_id, username)
        )
        if not await cursor.fetchone():
            raise HTTPException(status_code=403, detail="No tienes permiso para ver estos datos.")

        cursor = await conn.execute(
            "SELECT * FROM stock_counts WHERE session_id = ? AND counted_location = ? ORDER BY timestamp DESC",
            (session_id, location_code)
        )
        counts = await cursor.fetchall()
        return [dict(row) for row in counts]

@app.post('/api/counts')
async def add_count(data: Count, username: str = Depends(login_required)):
    print(f"Recibido: item={data.item_code}, cantidad={data.quantity}, ubicacion={data.location}")
    # La lógica de base de datos MySQL está comentada en el código original.
    # Si se necesita, se debe implementar aquí.
    return JSONResponse({'message': 'Endpoint de conteo no implementado en esta versión.'}, status_code=501)

@app.get('/api/find_item/{item_code}/{import_reference}')
async def find_item(item_code: str, import_reference: str, username: str = Depends(login_required)):
    item_details = await get_item_details_from_master_csv(item_code)
    if item_details is None:
        raise HTTPException(status_code=404, detail=f"Artículo {item_code} no encontrado en el maestro.")

    expected_quantity = await get_total_expected_quantity_for_item(item_code)
    original_bin = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin = await get_latest_relocated_bin_async(item_code)
    effective_bin_location = latest_relocated_bin if latest_relocated_bin else original_bin

    response_data = {
        "itemCode": item_details.get('Item_Code', item_code),
        "description": item_details.get('Item_Description', 'N/A'),
        "binLocation": effective_bin_location,
        "aditionalBins": item_details.get('Aditional_Bin_Location', 'N/A'),
        "weight": item_details.get('Weight_per_Unit', 'N/A'),
        "defaultQtyGrn": expected_quantity
    }
    return JSONResponse(content=response_data)

@app.post('/api/add_log')
async def add_log(data: LogEntry, username: str = Depends(login_required)):
    if data.quantity <= 0:
        raise HTTPException(status_code=400, detail="Cantidad debe ser > 0")

    item_details = await get_item_details_from_master_csv(data.itemCode)
    if item_details is None:
        raise HTTPException(status_code=400, detail=f"Artículo {data.itemCode} no existe.")

    original_bin_from_master = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin_for_item = await get_latest_relocated_bin_async(data.itemCode)
    bin_to_log_as_original = latest_relocated_bin_for_item if latest_relocated_bin_for_item else original_bin_from_master

    item_description = item_details.get('Item_Description', 'N/A')

    expected_quantity_grn = await get_total_expected_quantity_for_item(data.itemCode)
    total_already_received = await get_total_received_for_import_reference_async(data.importReference, data.itemCode)
    new_cumulative_total_received = total_already_received + data.quantity
    cumulative_difference = new_cumulative_total_received - expected_quantity_grn

    log_entry_data_for_db = {
        "timestamp": datetime.datetime.now().isoformat(timespec='seconds'),
        "importReference": data.importReference,
        "waybill": data.waybill,
        "itemCode": data.itemCode,
        "itemDescription": item_description,
        "binLocation": bin_to_log_as_original,
        "relocatedBin": data.relocatedBin,
        "qtyReceived": data.quantity,
        "qtyGrn": expected_quantity_grn,
        "difference": cumulative_difference
    }

    new_log_id = await save_log_entry_db_async(log_entry_data_for_db)

    if new_log_id is not None:
        log_entry_data_for_response = { "id": new_log_id, **log_entry_data_for_db }
        return JSONResponse(content={"message": "Registro añadido con éxito", "entry": log_entry_data_for_response}, status_code=201)
    else:
        raise HTTPException(status_code=500, detail="Error interno al guardar registro")

@app.put('/api/update_log/{log_id}')
async def update_log(log_id: int, data: dict, username: str = Depends(login_required)):
    try:
        qty_received_updated = int(data['qtyReceived'])
        if qty_received_updated < 0:
            raise HTTPException(status_code=400, detail="Cantidad recibida no puede ser negativa")
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Cantidad recibida debe ser un número válido")

    original_log_entry = await get_log_entry_by_id_async(log_id)
    if not original_log_entry:
        raise HTTPException(status_code=404, detail=f"Registro de log con ID {log_id} no encontrado.")

    import_reference = original_log_entry['importReference']
    item_code = original_log_entry['itemCode']
    expected_quantity_grn = await get_total_expected_quantity_for_item(item_code)
    total_already_received = await get_total_received_for_import_reference_async(import_reference, item_code)
    total_received_without_current_entry = total_already_received - original_log_entry.get('qtyReceived', 0)
    new_cumulative_total_received = total_received_without_current_entry + qty_received_updated
    updated_cumulative_difference = new_cumulative_total_received - expected_quantity_grn

    updated_log_data_for_db = {
        "waybill": data['waybill'],
        "relocatedBin": data['relocatedBin'],
        "qtyReceived": qty_received_updated,
        "difference": updated_cumulative_difference,
        "timestamp": datetime.datetime.now().isoformat(timespec='seconds'),
    }

    if await update_log_entry_db_async(log_id, updated_log_data_for_db):
        full_updated_entry_for_response = {**original_log_entry, **updated_log_data_for_db}
        full_updated_entry_for_response['qtyGrn'] = expected_quantity_grn
        return JSONResponse(content={"message": "Registro actualizado con éxito", "entry": full_updated_entry_for_response})
    else:
        raise HTTPException(status_code=500, detail="Error interno al actualizar el registro en BD")

@app.get('/api/get_logs')
async def get_logs(username: str = Depends(login_required)):
    logs = await load_log_data_db_async()
    return JSONResponse(content=logs)

@app.get('/api/stock')
async def get_stock(): # <--- Se eliminó la dependencia 'Depends(login_required)'
    stock_df = await get_stock_data()
    if stock_df is not None:
        stock_df_filled = stock_df.replace({np.nan: None})
        stock_records = stock_df_filled.to_dict(orient='records')
        return JSONResponse(content=stock_records)
    else:
        raise HTTPException(status_code=500, detail="No se pudieron obtener los datos de stock")

@app.get('/api/export_log')
async def export_log(username: str = Depends(login_required)):
    logs_data = await load_log_data_db_async()
    if not logs_data:
        raise HTTPException(status_code=404, detail="No hay registros para exportar")

    df = pd.DataFrame(logs_data)
    df_export = df[[
        'timestamp', 'importReference', 'waybill', 'itemCode', 'itemDescription',
        'binLocation', 'relocatedBin', 'qtyReceived', 'qtyGrn', 'difference'
    ]].rename(columns={
        'timestamp': 'Timestamp', 'importReference': 'Import Reference', 'waybill': 'Waybill',
        'itemCode': 'Item Code', 'itemDescription': 'Item Description',
        'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
        'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. Expected (Total)', 'difference': 'Difference'
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='InboundLogCompleto')
        worksheet = writer.sheets['InboundLogCompleto']
        for i, col_name in enumerate(df_export.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len

    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inbound_log_completo_{timestamp_str}.xlsx"
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.delete('/api/delete_log/{log_id}')
async def delete_log_api(log_id: int, username: str = Depends(login_required)):
    if await delete_log_entry_db_async(log_id):
        return JSONResponse(content={"message": f"Registro ID {log_id} eliminado con éxito."})
    else:
        raise HTTPException(status_code=404, detail=f"No se pudo eliminar el registro ID {log_id}. Es posible que ya haya sido eliminado.")

@app.post('/clear_database')
async def clear_database(password: str = Form(...)):
    if password != UPDATE_PASSWORD:
        # ARREGLADO: URL manual
        query_string = urlencode({'error': 'Contraseña incorrecta'})
        return RedirectResponse(url=f'/update?{query_string}', status_code=status.HTTP_302_FOUND)
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            await conn.execute('DELETE FROM logs')
            await conn.execute('DELETE FROM sqlite_sequence WHERE name="logs"')
            await conn.commit()
        # ARREGLADO: URL manual
        query_string = urlencode({'message': 'Base de datos limpiada'})
        return RedirectResponse(url=f'/update?{query_string}', status_code=status.HTTP_302_FOUND)
    except aiosqlite.Error as e:
        # ARREGLADO: URL manual
        query_string = urlencode({'error': str(e)})
        return RedirectResponse(url=f'/update?{query_string}', status_code=status.HTTP_302_FOUND)

@app.get('/api/stock_item/{item_code}')
async def stock_item(item_code: str):
    details = await get_item_details_from_master_csv(item_code)
    if details:
        response_data = {
            'Item_Code': details.get('Item_Code'),
            'Item_Description': details.get('Item_Description'),
            'Physical_Qty': details.get('Physical_Qty'),
            'Bin_1': details.get('Bin_1'),
            'Aditional_Bin_Location': details.get('Aditional_Bin_Location'),
            'Frozen_Qty': details.get('Frozen_Qty'),
            'ABC_Code_stockroom': details.get('ABC_Code_stockroom'),
            'Weight_per_Unit': details.get('Weight_per_Unit'),
            'SupersededBy': details.get('SupersededBy')
        }
        return JSONResponse(content=response_data)
    else:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")

@app.get('/api/get_item_details/{item_code}')
async def get_item_details_for_label(item_code: str, username: str = Depends(login_required)):
    details = await get_item_details_from_master_csv(item_code)
    if details:
        response_data = {
            'item_code': details.get('Item_Code'),
            'description': details.get('Item_Description'),
            'bin_location': details.get('Bin_1'),
            'additional_bins': details.get('Aditional_Bin_Location'),
            'weight_kg': details.get('Weight_per_Unit')
        }
        return JSONResponse(content=response_data)
    else:
        raise HTTPException(status_code=404, detail="Artículo no encontrado")


@app.get('/api/get_item_for_counting/{item_code}')
async def get_item_for_counting(item_code: str, username: str = Depends(login_required)):
    
    current_stage = 1 # Default
    
    # 1. Obtener la sesión activa del usuario y su etapa
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute(
            "SELECT inventory_stage FROM count_sessions WHERE user_username = ? AND status = 'in_progress' ORDER BY start_time DESC LIMIT 1",
            (username,)
        )
        active_session = await cursor.fetchone()
    
    if not active_session:
        raise HTTPException(status_code=403, detail="No tienes una sesión de conteo activa. Inicia una nueva sesión.")
        
    current_stage = active_session['inventory_stage']
    print(f"User {username} counting item {item_code} for stage {current_stage}")

    # 2. Aplicar lógica de etapa de reconteo (si aplica)
    if current_stage > 1:
        # Es un reconteo, verificar si el item está en la lista de tareas
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            cursor_recount = await conn.execute(
                "SELECT 1 FROM recount_list WHERE item_code = ? AND stage_to_count = ?",
                (item_code, current_stage)
            )
            item_in_list = await cursor_recount.fetchone()
            
        if not item_in_list:
            print(f"RECHAZADO: Item {item_code} no está en la lista de reconteo para la etapa {current_stage}.")
            raise HTTPException(status_code=404, detail=f"Item no requerido. Este item no está en la lista de reconteo para la Etapa {current_stage}.")

    # 3. Obtener detalles del item del maestro
    details = await get_item_details_from_master_csv(item_code)
    if details:
        # Si el item existe, devolver sus datos para el conteo ciego.
        response_data = {
            'item_code': details.get('Item_Code'),
            'description': details.get('Item_Description'),
            'bin_location': details.get('Bin_1')
        }
        return JSONResponse(content=response_data)
    else:
        # --- INICIO DE LA MODIFICACIÓN ---
        # Si el item no se encuentra en el maestro...
        if current_stage == 1:
            # ... y estamos en Etapa 1, permitimos el conteo ciego de items no registrados.
            print(f"Item {item_code} no encontrado en el maestro. Permitiendo conteo ciego para Etapa 1.")
            response_data = {
                'item_code': item_code,
                'description': 'ITEM NO ENCONTRADO',
                'bin_location': 'N/A'
            }
            return JSONResponse(content=response_data)
        else:
            # ... y estamos en una etapa de reconteo, un item no encontrado es un error.
            raise HTTPException(status_code=404, detail="Artículo no encontrado en el maestro de items.")
        # --- FIN DE LA MODIFICACIÓN ---

@app.post('/api/save_count')
async def save_count(data: StockCount, username: str = Depends(login_required)):
    """Guarda un conteo de stock, verificando la sesión, la etapa y la ubicación."""
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            
            # 1. Verificar que la sesión está activa y obtener su etapa
            cursor = await conn.execute(
                "SELECT id, inventory_stage FROM count_sessions WHERE id = ? AND user_username = ? AND status = 'in_progress'",
                (data.session_id, username)
            )
            active_session = await cursor.fetchone()
            
            if not active_session:
                raise HTTPException(status_code=403, detail="La sesión de conteo no es válida, está cerrada o no te pertenece.")
            
            current_stage = active_session['inventory_stage']
            print(f"User {username} saving item {data.item_code} for stage {current_stage}")

            # --- NUEVA VALIDACIÓN DE ETAPA ---
            if current_stage > 1:
                # Es un reconteo, verificar si el item está en la lista de tareas
                cursor_recount = await conn.execute(
                    "SELECT 1 FROM recount_list WHERE item_code = ? AND stage_to_count = ?",
                    (data.item_code, current_stage)
                )
                item_in_list = await cursor_recount.fetchone()
                
                if not item_in_list:
                    print(f"RECHAZADO (SAVE): Item {data.item_code} no está en la lista de reconteo para la etapa {current_stage}.")
                    raise HTTPException(status_code=400, detail=f"Item no requerido. Este item no está en la lista de reconteo (Etapa {current_stage}).")
            # --- FIN NUEVA VALIDACIÓN ---

            # 2. (Original) Verificar que la ubicación no esté cerrada para esta sesión
            cursor_loc = await conn.execute(
                "SELECT status FROM session_locations WHERE session_id = ? AND location_code = ?",
                (data.session_id, data.counted_location)
            )
            location_status = await cursor_loc.fetchone()
            if location_status and location_status['status'] == 'closed':
                raise HTTPException(status_code=400, detail=f"La ubicación {data.counted_location} ya está cerrada y no se puede modificar.")

            # 3. (Original) Insertar el conteo
            counted_qty = int(data.counted_qty)
            await conn.execute(
                '''
                INSERT INTO stock_counts (session_id, timestamp, item_code, item_description, counted_qty, counted_location, bin_location_system, username)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    data.session_id, datetime.datetime.now().isoformat(timespec='seconds'), data.item_code,
                    data.description, counted_qty, data.counted_location, data.bin_location_system, username
                )
            )
            
            # (Quitamos la lógica de UPDATE recount_list, eso lo hará el admin al cerrar la etapa)
            
            await conn.commit()
        return JSONResponse(content={'message': 'Conteo guardado con éxito'}, status_code=201)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Datos numéricos inválidos")
    except aiosqlite.Error as e:
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {e}")

@app.delete("/api/counts/{count_id}", status_code=status.HTTP_200_OK)
async def delete_stock_count(count_id: int, username: str = Depends(login_required)):
    """Elimina un registro de conteo de stock específico."""
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        # Opcional: Verificar que el conteo pertenece a una sesión del usuario.
        # Esta lógica puede ser más compleja si se requiere una validación estricta.
        cursor = await conn.execute("DELETE FROM stock_counts WHERE id = ?", (count_id,))
        await conn.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Registro de conteo no encontrado.")
        return {"message": "Registro de conteo eliminado con éxito."}

@app.get('/api/export_reconciliation')
async def export_reconciliation(username: str = Depends(login_required)):
    try:
        async with async_engine.connect() as conn:
            logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query('SELECT * FROM logs', sync_conn))

        grn_df = df_grn_cache

        if logs_df.empty or grn_df is None:
            raise HTTPException(status_code=404, detail="No hay datos suficientes para generar la conciliación")

        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)
        grn_df['Quantity'] = pd.to_numeric(grn_df['Quantity'], errors='coerce').fillna(0)

        item_totals = logs_df.groupby(['itemCode'])['qtyReceived'].sum().reset_index()
        item_totals = item_totals.rename(columns={'itemCode': 'Item_Code', 'qtyReceived': 'Total_Recibido'})

        grn_totals = grn_df.groupby(['GRN_Number', 'Item_Code', 'Item_Description'])['Quantity'].sum().reset_index()
        grn_totals = grn_totals.rename(columns={'Quantity': 'Total_Esperado'})

        merged_df = pd.merge(grn_totals, item_totals, on='Item_Code', how='outer')

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].fillna(0)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].fillna(0)
        merged_df['Diferencia'] = merged_df['Total_Recibido'] - merged_df['Total_Esperado']

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].astype(int)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].astype(int)
        merged_df['Diferencia'] = merged_df['Diferencia'].astype(int)

        df_for_export = merged_df.rename(columns={
            'GRN_Number': 'GRN',
            'Item_Code': 'Código de Ítem',
            'Item_Description': 'Descripción',
            'Total_Esperado': 'Cant. Esperada',
            'Total_Recibido': 'Cant. Recibida',
            'Diferencia': 'Diferencia'
        })

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_for_export.to_excel(writer, index=False, sheet_name='ReporteDeConciliacion')
            worksheet = writer.sheets['ReporteDeConciliacion']
            for i, col_name in enumerate(df_for_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_for_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_conciliacion_{timestamp_str}.xlsx"
        return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error interno al generar el archivo de conciliación: {e}")


@app.get('/api/debug/last_counts')
async def debug_last_counts(limit: int = 20, username: str = Depends(login_required)):
    """Endpoint de diagnóstico (temporal): devuelve los últimos `limit` registros de stock_counts.

    Protegido por login. Usar solo en desarrollo para verificar persistencia de conteos.
    """
    try:
        all_counts = await load_all_counts_db_async()
        # all_counts ya viene ordenado DESC por id en la consulta SQL
        return JSONResponse(content=all_counts[:int(limit)])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al leer conteos: {e}")

# --- Rutas para servir las páginas HTML ---
@app.get('/', response_class=HTMLResponse)
def home_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username # Return the redirect response if login fails
    return templates.TemplateResponse("inicio.html", {"request": request})

@app.get('/inbound', response_class=HTMLResponse)
def inbound_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username # Return the redirect response if login fails
    return templates.TemplateResponse("inbound.html", {"request": request})

@app.get('/update', response_class=HTMLResponse)
def update_files_get(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username
    return templates.TemplateResponse("update.html", {"request": request, "error": request.query_params.get('error'), "message": request.query_params.get('message')})

@app.post('/update', response_class=JSONResponse)
async def update_files_post(request: Request, item_master: UploadFile = File(None), grn_file: UploadFile = File(None), picking_file: UploadFile = File(None), username: str = Depends(login_required)):
    if not isinstance(username, str):
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"error": "Unauthorized"})
    
    files_uploaded = False
    message = ""
    error = ""

    if item_master and item_master.filename:
        if item_master.filename == os.path.basename(ITEM_MASTER_CSV_PATH):
            with open(ITEM_MASTER_CSV_PATH, "wb") as buffer:
                shutil.copyfileobj(item_master.file, buffer)
            message += f'Archivo "{item_master.filename}" actualizado. '
            files_uploaded = True
        else:
            error += f'Nombre incorrecto para maestro de items. Se esperaba "{os.path.basename(ITEM_MASTER_CSV_PATH)}". '

    if grn_file and grn_file.filename:
        if grn_file.filename == os.path.basename(GRN_CSV_FILE_PATH):
            with open(GRN_CSV_FILE_PATH, "wb") as buffer:
                shutil.copyfileobj(grn_file.file, buffer)
            message += f'Archivo "{grn_file.filename}" actualizado. '
            files_uploaded = True
        else:
            error += f'Nombre incorrecto para archivo GRN. Se esperaba "{os.path.basename(GRN_CSV_FILE_PATH)}". '

    if picking_file and picking_file.filename:
        if picking_file.filename == 'AURRSGLBD0240 - Unconfirmed Picking Notes.csv':
            with open(os.path.join(DATABASE_FOLDER, 'AURRSGLBD0240 - Unconfirmed Picking Notes.csv'), "wb") as buffer:
                shutil.copyfileobj(picking_file.file, buffer)
            message += f'Archivo "{picking_file.filename}" actualizado. '
            files_uploaded = True
        else:
            error += f'Nombre incorrecto para archivo de picking. Se esperaba "AURRSGLBD0240 - Unconfirmed Picking Notes.csv". '

    if files_uploaded:
        await load_csv_data()

    if not files_uploaded and not error:
        error = "No seleccionaste ningún archivo para subir."

    if error:
        return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={"error": error})
    
    return JSONResponse(content={"message": message})

@app.get('/view_logs', response_class=HTMLResponse)
async def view_logs(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username
    all_logs = await load_log_data_db_async()
    return templates.TemplateResponse('view_logs.html', {"request": request, "logs": all_logs})

@app.get('/label', response_class=HTMLResponse)
def label_page(request: Request):
    return templates.TemplateResponse('label.html', {"request": request})

@app.get('/counts', response_class=HTMLResponse)
def counts_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username
    return templates.TemplateResponse('counts.html', {"request": request})

@app.get('/stock', response_class=HTMLResponse)
async def stock_page(request: Request): # <--- Se eliminó la dependencia 'Depends(login_required)'
    # También se eliminó el bloque que verificaba al usuario
    # if not isinstance(username, str):
    #     return username
    return templates.TemplateResponse('stock.html', {"request": request})

@app.get("/api/picking/order/{order_number}/{despatch_number}")
async def get_picking_order(order_number: str, despatch_number: str):
    DB_FOLDER_PATH = os.path.join(os.path.dirname(__file__), 'databases')
    try:
        picking_file_path = os.path.join(DB_FOLDER_PATH, "AURRSGLBD0240 - Unconfirmed Picking Notes.csv")
        if not os.path.exists(picking_file_path):
            raise HTTPException(status_code=404, detail="El archivo de picking (AURRSGLBD0240.csv) no se encuentra.")

        df = pd.read_csv(picking_file_path, dtype=str)
        
        # Asegurarse de que las columnas existen
        required_columns = ["ORDER_", "DESPATCH_", "ITEM", "DESCRIPTION", "QTY", "CUSTOMER_NAME"]
        if not all(col in df.columns for col in required_columns):
            raise HTTPException(status_code=500, detail="El archivo CSV no tiene las columnas esperadas.")

        # Filtrar los datos
        order_data = df[
            (df["ORDER_"] == order_number) & 
            (df["DESPATCH_"] == despatch_number)
        ]

        if order_data.empty:
            raise HTTPException(status_code=404, detail="Pedido no encontrado.")

        # Renombrar las columnas para que coincidan con el frontend
        order_data = order_data.rename(columns={
            "ORDER_": "Order Number",
            "DESPATCH_": "Despatch Number",
            "ITEM": "Item Code",
            "DESCRIPTION": "Item Description",
            "QTY": "Qty",
            "CUSTOMER_NAME": "Customer Name"
        })

        # Reemplazar NaN con None para que sea compatible con JSON
        order_data = order_data.where(pd.notnull(order_data), None)

        return JSONResponse(content=order_data.to_dict(orient="records"))

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/picking', response_class=HTMLResponse)
async def picking_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username
    return templates.TemplateResponse("picking.html", {"request": request, "username": username})

@app.post('/api/save_picking_audit')
async def save_picking_audit(audit_data: PickingAudit, username: str = Depends(login_required)):
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        try:
            # 1. Insertar la auditoría principal
            cursor = await conn.execute(
                '''
                INSERT INTO picking_audits (order_number, despatch_number, customer_name, username, timestamp, status)
                VALUES (?, ?, ?, ?, ?, ?)
                ''',
                (
                    audit_data.order_number,
                    audit_data.despatch_number,
                    audit_data.customer_name,
                    username,
                    datetime.datetime.now().isoformat(timespec='seconds'),
                    audit_data.status
                )
            )
            await conn.commit()
            audit_id = cursor.lastrowid

            # 2. Insertar los items de la auditoría
            items_to_insert = []
            for item in audit_data.items:
                difference = item.qty_scan - item.qty_req
                items_to_insert.append((
                    audit_id,
                    item.code,
                    item.description,
                    item.qty_req,
                    item.qty_scan,
                    difference
                ))

            await conn.executemany(
                '''
                INSERT INTO picking_audit_items (audit_id, item_code, description, qty_req, qty_scan, difference)
                VALUES (?, ?, ?, ?, ?, ?)
                ''',
                items_to_insert
            )
            await conn.commit()

            return JSONResponse(content={"message": "Auditoría de picking guardada con éxito", "audit_id": audit_id}, status_code=201)

        except aiosqlite.Error as e:
            print(f"Database error in save_picking_audit: {e}")
            raise HTTPException(status_code=500, detail=f"Error de base de datos: {e}")

@app.get('/view_picking_audits', response_class=HTMLResponse)
async def view_picking_audits_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username

    audits = await load_picking_audits_from_db()
    return templates.TemplateResponse('view_picking_audits.html', {"request": request, "audits": audits})

async def load_picking_audits_from_db():
    audits = []
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT * FROM picking_audits ORDER BY id DESC")
        rows = await cursor.fetchall()
        for row in rows:
            audit = dict(row)
            cursor_items = await conn.execute("SELECT * FROM picking_audit_items WHERE audit_id = ?", (audit['id'],))
            items = await cursor_items.fetchall()
            audit['items'] = [dict(item) for item in items]
            audits.append(audit)
    return audits

@app.get('/view_counts', response_class=HTMLResponse)
async def view_counts_page(request: Request, username: str = Depends(login_required)):
    async with async_engine.connect() as conn:
        all_counts = await conn.run_sync(
            lambda sync_conn: pd.read_sql_query(
                """
                SELECT
                    sc.*,
                    cs.inventory_stage  -- --- NUEVO: Traer la etapa de la sesión ---
                FROM
                    stock_counts sc
                JOIN 
                    count_sessions cs ON sc.session_id = cs.id
                ORDER BY
                    sc.id DESC
                """,
                sync_conn
            )
        )
    
    # --- LÓGICA MODIFICADA PARA AGRUPAR POR ETAPA ---
    # Convertir a dicts
    all_counts_list = all_counts.to_dict(orient='records')
    
    # Agrupar conteos por item, ubicación Y etapa
    grouped_counts = {}
    for count in all_counts_list:
        key = (count['item_code'], count['counted_location'], count['inventory_stage'])
        if key not in grouped_counts:
            grouped_counts[key] = {
                **count,
                'counted_qty': 0
            }
        grouped_counts[key]['counted_qty'] += count['counted_qty']
        
    # Usar los conteos agrupados
    all_counts_list = list(grouped_counts.values())

    master_map = master_qty_map
    session_map = {}
    session_ids = list({c.get('session_id') for c in all_counts_list if c.get('session_id') is not None})
    if session_ids:
        try:
            async with aiosqlite.connect(DB_FILE_PATH) as conn:
                conn.row_factory = aiosqlite.Row
                placeholders = ','.join('?' * len(session_ids))
                query = f"SELECT id, user_username FROM count_sessions WHERE id IN ({placeholders})"
                async with conn.execute(query, tuple(session_ids)) as cursor:
                    rows = await cursor.fetchall()
                    for r in rows:
                        session_map[r['id']] = r['user_username']
        except Exception:
            session_map = {}

    enriched_counts = []
    for count in all_counts_list:
        item_code = count.get('item_code')
        system_qty = None
        raw_system = master_map.get(item_code) if master_map else None
        if raw_system not in (None, ''):
            try:
                system_qty = int(float(raw_system))
            except (ValueError, TypeError):
                system_qty = None

        try:
            counted_qty = int(count.get('counted_qty') or 0)
        except (ValueError, TypeError):
            counted_qty = 0

        difference = (counted_qty - system_qty) if system_qty is not None else None

        enriched = dict(count)
        enriched['system_qty'] = system_qty
        enriched['difference'] = difference
        enriched['username'] = count.get('username') or (session_map.get(count.get('session_id')) if session_map else None)
        enriched_counts.append(enriched)

    usernames = sorted({u for u in session_map.values() if u})

    return templates.TemplateResponse('view_counts.html', {"request": request, "counts": enriched_counts, "usernames": usernames})


@app.get('/api/export_counts')
async def export_counts(username: str = Depends(login_required)):
    """Exporta todos los conteos enriquecidos a Excel (incluye usuario, system_qty y diferencia)."""
    
    # --- NUEVO: Traer la etapa de inventario en la consulta ---
    async with async_engine.connect() as conn:
        all_counts_df = await conn.run_sync(
            lambda sync_conn: pd.read_sql_query(
                """
                SELECT 
                    sc.*, 
                    cs.inventory_stage 
                FROM 
                    stock_counts sc
                JOIN 
                    count_sessions cs ON sc.session_id = cs.id
                ORDER BY sc.id DESC
                """, 
                sync_conn
            )
        )
    all_counts = all_counts_df.to_dict(orient='records')


    master_map = master_qty_map
    session_map = {}
    session_ids = list({c.get('session_id') for c in all_counts if c.get('session_id') is not None})
    if session_ids:
        try:
            async with aiosqlite.connect(DB_FILE_PATH) as conn:
                conn.row_factory = aiosqlite.Row
                placeholders = ','.join('?' * len(session_ids))
                query = f"SELECT id, user_username FROM count_sessions WHERE id IN ({placeholders})"
                async with conn.execute(query, tuple(session_ids)) as cursor:
                    rows = await cursor.fetchall()
                    for r in rows:
                        session_map[r['id']] = r['user_username']
        except Exception:
            session_map = {}

    enriched_rows = []
    for count in all_counts:
        item_code = count.get('item_code')
        raw_system = master_map.get(item_code) if master_map else None
        system_qty = None
        if raw_system not in (None, ''):
            try:
                system_qty = int(float(raw_system))
            except (ValueError, TypeError):
                system_qty = None

        try:
            counted_qty = int(count.get('counted_qty') or 0)
        except (ValueError, TypeError):
            counted_qty = 0

        difference = (counted_qty - system_qty) if system_qty is not None else None

        enriched = {
            'id': count.get('id'),
            'session_id': count.get('session_id'),
            'inventory_stage': count.get('inventory_stage'), # --- NUEVO: Añadir etapa al reporte ---
            'username': count.get('username') or (session_map.get(count.get('session_id')) if session_map else None),
            'timestamp': count.get('timestamp'),
            'item_code': item_code,
            'item_description': count.get('item_description'),
            'counted_location': count.get('counted_location'),
            'counted_qty': counted_qty,
            'system_qty': system_qty,
            'difference': difference,
            'bin_location_system': count.get('bin_location_system')
        }
        enriched_rows.append(enriched)

    # Construir DataFrame y exportar a Excel
    df = pd.DataFrame(enriched_rows)
    # Reordenar columnas para la exportación
    columns_order = ['id', 'session_id', 'inventory_stage', 'username', 'timestamp', 'item_code', 'item_description', 'counted_location', 'counted_qty', 'system_qty', 'difference', 'bin_location_system']
    df = df[columns_order]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Conteos')
        worksheet = writer.sheets['Conteos']
        for i, col_name in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len

    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"conteos_export_{timestamp_str}.xlsx"
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})





@app.get('/api/counts/stats')
async def get_count_stats(username: str = Depends(login_required)):
    """Devuelve estadísticas sobre los conteos de stock."""
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            
            # --- NUEVO: Obtener etapa actual ---
            cursor = await conn.execute("SELECT value FROM app_state WHERE key = 'current_inventory_stage'")
            stage_row = await cursor.fetchone()
            current_stage = int(stage_row['value']) if stage_row else 1
            
            # --- NUEVO: Filtrar consultas por la etapa actual ---
            
            # 1. Total de ubicaciones contadas (en esta etapa)
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT sc.counted_location) 
                FROM stock_counts sc
                JOIN count_sessions cs ON sc.session_id = cs.id
                WHERE cs.inventory_stage = ?
                """, (current_stage,)
            )
            counted_locations = (await cursor.fetchone())[0]

            # 2. Total de items contados (grupos de item/ubicación) (en esta etapa)
            cursor = await conn.execute(
                """
                SELECT COUNT(*) FROM (
                    SELECT DISTINCT sc.item_code, sc.counted_location 
                    FROM stock_counts sc
                    JOIN count_sessions cs ON sc.session_id = cs.id
                    WHERE cs.inventory_stage = ?
                )
                """, (current_stage,)
            )
            total_items_counted = (await cursor.fetchone())[0]

            # 3. Total de items con stock (del maestro de items)
            total_items_with_stock = 0
            if master_qty_map:
                for qty in master_qty_map.values():
                    if qty is not None and qty > 0:
                        total_items_with_stock += 1
            
            # 4. Items con diferencias (en esta etapa)
            cursor = await conn.execute(
                """
                SELECT sc.item_code, sc.counted_qty 
                FROM stock_counts sc
                JOIN count_sessions cs ON sc.session_id = cs.id
                WHERE cs.inventory_stage = ?
                """, (current_stage,)
            )
            all_counts_stage = await cursor.fetchall()
            
            items_with_differences = 0
            processed_items_stage = set()

            for count in all_counts_stage:
                item_code = count['item_code']
                if item_code not in processed_items_stage:
                    system_qty_raw = master_qty_map.get(item_code)
                    system_qty = 0
                    if system_qty_raw is not None:
                        try:
                            system_qty = int(float(system_qty_raw))
                        except (ValueError, TypeError):
                            system_qty = 0
                    
                    # Sumar todos los conteos para este item_code (SOLO en esta etapa)
                    cursor = await conn.execute(
                        """
                        SELECT SUM(sc.counted_qty) 
                        FROM stock_counts sc
                        JOIN count_sessions cs ON sc.session_id = cs.id
                        WHERE sc.item_code = ? AND cs.inventory_stage = ?
                        """, (item_code, current_stage)
                    )
                    total_counted_for_item = (await cursor.fetchone())[0]

                    if total_counted_for_item != system_qty:
                        items_with_differences += 1
                    
                    processed_items_stage.add(item_code)

            # 5. Ubicaciones con stock (contadas en esta etapa)
            cursor = await conn.execute(
                """
                SELECT COUNT(DISTINCT sc.counted_location) 
                FROM stock_counts sc
                JOIN count_sessions cs ON sc.session_id = cs.id
                WHERE cs.inventory_stage = ? AND sc.counted_qty > 0
                """, (current_stage,)
            )
            locations_with_stock = (await cursor.fetchone())[0]

            # 6. Total de ubicaciones con stock (del maestro de items)
            total_locations_with_stock = 0
            if df_master_cache is not None:
                stock_items = df_master_cache[pd.to_numeric(df_master_cache['Physical_Qty'], errors='coerce').fillna(0) > 0]
                bin_1_locations = stock_items['Bin_1'].dropna().unique()
                total_locations_with_stock = len(bin_1_locations)

            # --- NUEVO: Añadir 'current_stage' a la respuesta ---
            return JSONResponse(content={
                "current_stage": current_stage, 
                "total_items_with_stock": total_items_with_stock,
                "counted_locations": counted_locations,
                "total_items_counted": total_items_counted,
                "items_with_differences": items_with_differences,
                "locations_with_stock": locations_with_stock,
                "total_locations_with_stock": total_locations_with_stock
            })

    except aiosqlite.Error as e:
        raise HTTPException(status_code=500, detail=f"Error de base de datos: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {e}")


@app.get('/reconciliation', response_class=HTMLResponse)
async def reconciliation_page(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username
    try:
        async with async_engine.connect() as conn:
            logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query('SELECT * FROM logs', sync_conn))

        grn_df = df_grn_cache

        if logs_df.empty or grn_df is None:
            return templates.TemplateResponse('reconciliation.html', {"request": request, "tables": []})

        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)
        grn_df['Quantity'] = pd.to_numeric(grn_df['Quantity'], errors='coerce').fillna(0)

        item_totals = logs_df.groupby(['itemCode'])['qtyReceived'].sum().reset_index()
        item_totals = item_totals.rename(columns={'itemCode': 'Item_Code', 'qtyReceived': 'Total_Recibido'})

        grn_totals = grn_df.groupby(['GRN_Number', 'Item_Code', 'Item_Description'])['Quantity'].sum().reset_index()
        grn_totals = grn_totals.rename(columns={'Quantity': 'Total_Esperado'})

        merged_df = pd.merge(grn_totals, item_totals, on='Item_Code', how='outer')

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].fillna(0)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].fillna(0)
        merged_df['Diferencia'] = merged_df['Total_Recibido'] - merged_df['Total_Esperado']

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].astype(int)
        merged_df['Total_Esperado'] = merged_df['Total_Esperado'].astype(int)
        merged_df['Diferencia'] = merged_df['Diferencia'].astype(int)

        merged_df = merged_df.rename(columns={
            'GRN_Number': 'GRN',
            'Item_Code': 'Código de Ítem',
            'Item_Description': 'Descripción',
            'Total_Esperado': 'Cant. Esperada',
            'Total_Recibido': 'Cant. Recibida',
            'Diferencia': 'Diferencia'
        })

        return templates.TemplateResponse('reconciliation.html', {
            "request": request,
            "tables": [merged_df.to_html(classes='min-w-full leading-normal', border=0, index=False)],
            "titles": merged_df.columns.values
        })

    except Exception as e:
        return templates.TemplateResponse('reconciliation.html', {"request": request, "error": str(e)})

@app.get('/register', response_class=HTMLResponse)
def register_get(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})

@app.post('/register', response_class=HTMLResponse)
async def register_post(request: Request, username: str = Form(...), password: str = Form(...)):
    if not username or not password:
        # En una app real, manejaríamos esto con mensajes de error en la plantilla
        return RedirectResponse(url=str(request.url.replace(query='error=Usuario y contraseña requeridos')), status_code=status.HTTP_302_FOUND)

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        cursor = await conn.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = await cursor.fetchone()

    if user:
        return RedirectResponse(url=str(request.url.replace(query='error=El usuario ya existe')), status_code=status.HTTP_302_FOUND)

    hashed_password = generate_password_hash(password)
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        await conn.execute("INSERT INTO users (username, password_hash) VALUES (?, ?)", (username, hashed_password))
        await conn.commit()

    # ARREGLADO: URL manual
    query_string = urlencode({'message': 'Registro exitoso. Pendiente de aprobación.'})
    return RedirectResponse(url=f'/login?{query_string}', status_code=status.HTTP_302_FOUND)

@app.get('/login', response_class=HTMLResponse, name='login')
def login_get(request: Request):
    if get_current_user(request):
        return RedirectResponse(url='/', status_code=status.HTTP_302_FOUND)
    return templates.TemplateResponse("login.html", {"request": request, "message": request.query_params.get("message"), "error": request.query_params.get("error")})

@app.post('/login')
async def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = await cursor.fetchone()

    if user and check_password_hash(user['password_hash'], password):
        if user['is_approved'] == 1:
            # Éxito: redirigimos al root y establecemos la cookie de sesión.
            # Anteriormente devolvíamos JSON, lo que hacía que el navegador mostrara
            # el JSON crudo en la página. Usamos RedirectResponse para evitar ese
            # comportamiento y no mostrar el mensaje de 'Login successful'.
            response = RedirectResponse(url='/', status_code=status.HTTP_302_FOUND)
            # Adjuntamos la cookie a la respuesta de redirección.
            # Usar secure sólo si la petición original fue HTTPS (evita problemas en desarrollo)
            response.set_cookie(key="username", value=username, httponly=True, samesite='lax', secure=(request.url.scheme == 'https'), path='/')
            return response
        else:
            # Error de aprobación: Enviamos un JSON con el error específico.
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Tu cuenta aún no ha sido aprobada."
            )
    else:
        # Error de credenciales: Enviamos un JSON con el error específico.
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos."
        )

@app.get('/logout')
def logout(request: Request):
    query_string = urlencode({'message': 'Has cerrado la sesión'})
    response = RedirectResponse(url=f'/login?{query_string}', status_code=status.HTTP_302_FOUND)
    response.delete_cookie("username")
    response.delete_cookie("admin_logged_in")
    return response

@app.get('/admin_inventory', response_class=RedirectResponse)
async def redirect_admin_inventory():
    return RedirectResponse(url='/admin/inventory')

def admin_login_required(request: Request):
    if not request.cookies.get("admin_logged_in"):
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)
    return True

@app.get('/api/export_recount_list/{stage_number}', name='export_recount_list')
async def export_recount_list(request: Request, stage_number: int, admin: bool = Depends(admin_login_required)):
    """Exporta la lista de items a recontar para una etapa específica."""
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute(
            "SELECT item_code FROM recount_list WHERE stage_to_count = ?",
            (stage_number,)
        )
        items_to_recount = await cursor.fetchall()

    if not items_to_recount:
        raise HTTPException(status_code=404, detail=f"No hay items en la lista de reconteo para la Etapa {stage_number}.")

    enriched_data = []
    for item in items_to_recount:
        item_code = item['item_code']
        details = await get_item_details_from_master_csv(item_code)
        if details:
            enriched_data.append({
                'Código de Item': item_code,
                'Descripción': details.get('Item_Description', 'N/A'),
                'Ubicación en Sistema': details.get('Bin_1', 'N/A')
            })
        else:
            # Para items "fantasma" que no están en el maestro
            enriched_data.append({
                'Código de Item': item_code,
                'Descripción': 'ITEM NO ENCONTRADO EN MAESTRO',
                'Ubicación en Sistema': 'N/A'
            })

    df = pd.DataFrame(enriched_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Reconteo_Etapa_{stage_number}')
        worksheet = writer.sheets[f'Reconteo_Etapa_{stage_number}']
        for i, col_name in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len
    
    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lista_reconteo_etapa_{stage_number}_{timestamp_str}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get('/admin/inventory', response_class=HTMLResponse, name='admin_inventory')
async def admin_inventory_get(request: Request, admin: bool = Depends(admin_login_required)):
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT * FROM app_state WHERE key = 'current_inventory_stage'")
        stage = await cursor.fetchone()
        if not stage:
            # Si no existe, inicializamos a etapa 0 (inactivo)
            await conn.execute("INSERT OR REPLACE INTO app_state (key, value) VALUES ('current_inventory_stage', '0')")
            await conn.commit()
            cursor = await conn.execute("SELECT * FROM app_state WHERE key = 'current_inventory_stage'")
            stage = await cursor.fetchone()

    message = request.query_params.get('message')
    error = request.query_params.get('error')
    
    return templates.TemplateResponse('admin_inventory.html', {
        "request": request, 
        "stage": stage,
        "message": message,
        "error": error
    })

@app.post('/admin/inventory/start_stage_1', name='start_inventory_stage_1')
async def start_inventory_stage_1(request: Request, admin: bool = Depends(admin_login_required)):
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)
    
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            # --- INICIO DE LA CORRECCIÓN ---
            # Limpiar todas las tablas relacionadas con el ciclo de conteo anterior
            print("Limpiando tablas de inventario para un nuevo ciclo...")
            await conn.execute('DELETE FROM stock_counts')
            await conn.execute('DELETE FROM count_sessions')
            await conn.execute('DELETE FROM session_locations')
            await conn.execute('DELETE FROM recount_list')
            
            # --- NUEVO: Reiniciar los contadores de autoincremento ---
            await conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('stock_counts', 'count_sessions', 'session_locations', 'recount_list')")
            print("Tablas de inventario y contadores de ID reiniciados.")
            # --- FIN DE LA CORRECCIÓN ---

            # Establecer el estado a Etapa 1
            await conn.execute("UPDATE app_state SET value = '1' WHERE key = 'current_inventory_stage'")
            await conn.commit()
        
        query_params = urlencode({"message": "Inventario reiniciado en Etapa 1. Todos los datos y contadores han sido reseteados."})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)
    except aiosqlite.Error as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)

@app.post('/admin/inventory/advance/{next_stage}', name='advance_inventory_stage')
async def advance_inventory_stage(request: Request, next_stage: int, admin: bool = Depends(admin_login_required)):
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)

    prev_stage = next_stage - 1
    
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            
            # 1. Obtener todos los conteos de la etapa anterior
            query = """
                SELECT 
                    sc.item_code, 
                    SUM(sc.counted_qty) as total_counted
                FROM 
                    stock_counts sc
                JOIN 
                    count_sessions cs ON sc.session_id = cs.id
                WHERE 
                    cs.inventory_stage = ?
                GROUP BY 
                    sc.item_code
            """
            cursor = await conn.execute(query, (prev_stage,))
            counted_items = await cursor.fetchall()
            
            # 2. Limpiar la lista de reconteo para la *próxima* etapa
            await conn.execute("DELETE FROM recount_list WHERE stage_to_count = ?", (next_stage,))

            # 3. Comparar con el maestro y generar la nueva lista de reconteo
            items_for_recount = []
            for item in counted_items:
                item_code = item['item_code']
                total_counted = item['total_counted']
                
                system_qty = master_qty_map.get(item_code)
                system_qty = int(system_qty) if system_qty is not None else 0

                if total_counted != system_qty:
                    items_for_recount.append((item_code, next_stage))

            # 4. Insertar los items en la lista de reconteo
            if items_for_recount:
                await conn.executemany(
                    "INSERT INTO recount_list (item_code, stage_to_count) VALUES (?, ?)",
                    items_for_recount
                )

            # 5. Actualizar el estado global a la nueva etapa
            await conn.execute("UPDATE app_state SET value = ? WHERE key = 'current_inventory_stage'", (str(next_stage),))
            
            await conn.commit()

        message = f"Proceso completado. Etapa de inventario avanzada a {next_stage}. Se encontraron {len(items_for_recount)} items con diferencias."
        query_params = urlencode({"message": message})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)

    except aiosqlite.Error as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)
    except Exception as e:
        query_params = urlencode({"error": f"Error inesperado: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)

@app.post('/admin/inventory/finalize', name='finalize_inventory')
async def finalize_inventory(request: Request, admin: bool = Depends(admin_login_required)):
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)
    
    try:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            # Cambiamos el estado a '0' para indicar que no hay un ciclo activo
            await conn.execute("UPDATE app_state SET value = '0' WHERE key = 'current_inventory_stage'")
            await conn.commit()
        
        query_params = urlencode({"message": "Ciclo de inventario finalizado y cerrado."})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)
    except aiosqlite.Error as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)



@app.get('/admin/login', response_class=HTMLResponse, name='admin_login_get')
def admin_login_get(request: Request):
    return templates.TemplateResponse("admin_login.html", {"request": request})

@app.post('/admin/login', response_class=HTMLResponse, name='admin_login_post')
def admin_login_post(request: Request, password: str = Form(...)):
    if password == UPDATE_PASSWORD:
        response = RedirectResponse(url=request.url_for('admin_inventory'), status_code=status.HTTP_302_FOUND)
        response.set_cookie(key="admin_logged_in", value="true", httponly=True, samesite='lax', secure=(request.url.scheme == 'https'))
        return response
    else:
        return templates.TemplateResponse("admin_login.html", {"request": request, "error": "Contraseña incorrecta"})

@app.get('/admin/users', response_class=HTMLResponse, name='admin_users_get')
async def admin_users_get(request: Request, admin: bool = Depends(admin_login_required)):
    if not admin:
        return RedirectResponse(url='/admin/login', status_code=status.HTTP_302_FOUND)
    
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT id, username, is_approved FROM users ORDER BY id DESC")
        users = await cursor.fetchall()
    
    reset_token = request.query_params.get('reset_token')
    reset_user = request.query_params.get('reset_user')
    return templates.TemplateResponse('admin_users.html', {"request": request, "users": users, "reset_token": reset_token, "reset_user": reset_user})

@app.post('/admin/check_password/{user_id}')
async def check_password(user_id: int, request: Request):
    if not request.cookies.get("admin_logged_in"):
        raise HTTPException(status_code=403, detail="No tienes permiso para realizar esta acción.")
    
    data = await request.json()
    password = data.get('password')

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT password_hash FROM users WHERE id = ?", (user_id,))
        user = await cursor.fetchone()

    if user and check_password_hash(user['password_hash'], password):
        return JSONResponse(content={"matches": True})
    else:
        return JSONResponse(content={"matches": False})

@app.post('/admin/approve/{user_id}')
async def approve_user(user_id: int, request: Request):
    if not request.cookies.get("admin_logged_in"):
        return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        await conn.execute("UPDATE users SET is_approved = 1 WHERE id = ?", (user_id,))
        await conn.commit()
    return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)

@app.post('/admin/delete/{user_id}')
async def delete_user(user_id: int, request: Request):
    if not request.cookies.get("admin_logged_in"):
        return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        await conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        await conn.commit()
    return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)

@app.post('/admin/reset_password/{user_id}')
async def reset_password(user_id: int, request: Request):
    if not request.cookies.get("admin_logged_in"):
        return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)
    
    # Obtener el username del usuario
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        cursor = await conn.execute("SELECT username FROM users WHERE id = ?", (user_id,))
        user = await cursor.fetchone()
    
    if not user:
        return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='error=Usuario no encontrado')), status_code=status.HTTP_302_FOUND)
    
    username = user[0]
    # Invalidar la contraseña actual (para que la antigua deje de funcionar)
    temp_invalid = secrets.token_urlsafe(32)
    hashed_invalid_for_user = generate_password_hash(temp_invalid)
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        await conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (hashed_invalid_for_user, user_id))
        await conn.commit()

    # Generar token de un solo uso y almacenarlo con expiración
    token = secrets.token_urlsafe(32)
    expires_at = (datetime.datetime.utcnow() + datetime.timedelta(hours=1)).isoformat()
    created_at = datetime.datetime.utcnow().isoformat()

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        await conn.execute(
            "INSERT INTO password_reset_tokens (user_id, token, expires_at, used, created_at) VALUES (?, ?, ?, 0, ?)",
            (user_id, token, expires_at, created_at)
        )
        await conn.commit()

    # Renderizar directamente la página admin con el token en contexto (NO poner token en la URL)
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT id, username, is_approved FROM users ORDER BY id DESC")
        users = await cursor.fetchall()

    return templates.TemplateResponse('admin_users.html', {"request": request, "users": users, "reset_token": token, "reset_user": username})


@app.get('/set_password')
async def set_password(request: Request, token: Optional[str] = None):
    """Muestra el formulario para cambiar la contraseña usando un token de un solo uso."""
    context = {"request": request, "token": "", "username": "", "error": None, "message": None}
    if token:
        async with aiosqlite.connect(DB_FILE_PATH) as conn:
            conn.row_factory = aiosqlite.Row
            cursor = await conn.execute("SELECT user_id, token, expires_at, used FROM password_reset_tokens WHERE token = ?", (token,))
            row = await cursor.fetchone()
            if not row:
                context['error'] = "Token inválido o inexistente."
                return templates.TemplateResponse("set_password.html", context)

            if row['used']:
                context['error'] = "Este token ya fue utilizado."
                return templates.TemplateResponse("set_password.html", context)

            expires_at = datetime.datetime.fromisoformat(row['expires_at'])
            if datetime.datetime.utcnow() > expires_at:
                context['error'] = "El token ha expirado. Solicita un nuevo restablecimiento."
                return templates.TemplateResponse("set_password.html", context)

            # Obtener username para mostrar en el formulario
            cursor = await conn.execute("SELECT username FROM users WHERE id = ?", (row['user_id'],))
            user_row = await cursor.fetchone()
            context['username'] = user_row['username'] if user_row else ''
            context['token'] = token

    return templates.TemplateResponse("set_password.html", context)


@app.post('/set_password')
async def set_password_post(request: Request, token: str = Form(...), new_password: str = Form(...), confirm_password: str = Form(...)):
    """Procesa el cambio de contraseña usando el token: valida token y actualiza la contraseña.

    Requisitos mínimos de contraseña: al menos 8 caracteres, al menos una letra y al menos un dígito.
    """

    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        conn.row_factory = aiosqlite.Row
        cursor = await conn.execute("SELECT id, user_id, expires_at, used FROM password_reset_tokens WHERE token = ?", (token,))
        token_row = await cursor.fetchone()
        if not token_row:
            return templates.TemplateResponse("set_password.html", {"request": request, "token": "", "username": "", "error": "Token inválido.", "message": None})

        if token_row['used']:
            return templates.TemplateResponse("set_password.html", {"request": request, "token": "", "username": "", "error": "Token ya utilizado.", "message": None})

        expires_at = datetime.datetime.fromisoformat(token_row['expires_at'])
        if datetime.datetime.utcnow() > expires_at:
            return templates.TemplateResponse("set_password.html", {"request": request, "token": "", "username": "", "error": "El token ha expirado.", "message": None})

        user_id = token_row['user_id']
        # Obtener username para volver a mostrar en caso de error
        cursor = await conn.execute("SELECT username FROM users WHERE id = ?", (user_id,))
        user_row = await cursor.fetchone()
        username = user_row['username'] if user_row else ''

    # Validaciones del lado servidor
    if new_password != confirm_password:
        return templates.TemplateResponse("set_password.html", {"request": request, "token": token, "username": username, "error": "Las contraseñas nuevas no coinciden.", "message": None})

    # Política mínima: longitud >=8, al menos una letra y un dígito
    if len(new_password) < 8 or not any(c.isalpha() for c in new_password) or not any(c.isdigit() for c in new_password):
        return templates.TemplateResponse("set_password.html", {"request": request, "token": token, "username": username, "error": "La contraseña debe tener al menos 8 caracteres, incluir letras y dígitos.", "message": None})

    # Si pasa validaciones, actualizar la contraseña y marcar token como usado
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        new_hash = generate_password_hash(new_password)
        await conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, user_id))
        await conn.execute("UPDATE password_reset_tokens SET used = 1 WHERE id = ?", (token_row['id'],))
        await conn.commit()

    # Redirigir al login con mensaje de éxito
    query_string = urlencode({'message': 'Contraseña actualizada con éxito. Por favor inicie sesión.'})
    return RedirectResponse(url=f'/login?{query_string}', status_code=status.HTTP_302_FOUND)

@app.post('/admin/reset_count_sessions/{user_id}', name='reset_count_sessions')
async def reset_count_sessions(user_id: int, request: Request):
    """
    MODIFICADO: Esta función ahora CIERRA FORZOSAMENTE la sesión activa del usuario,
    preservando los datos del conteo, en lugar de borrarlos.
    """
    if not request.cookies.get("admin_logged_in"):
        return RedirectResponse(url=str(request.url.replace(path='/admin/users', query='')), status_code=status.HTTP_302_FOUND)
    
    async with aiosqlite.connect(DB_FILE_PATH) as conn:
        cursor = await conn.execute("SELECT username FROM users WHERE id = ?", (user_id,))
        user = await cursor.fetchone()
        
        if user:
            username = user[0]
            now = datetime.datetime.now().isoformat(timespec='seconds')
            
            # Cierra (Completa) la sesión activa
            await conn.execute(
                "UPDATE count_sessions SET status = 'completed', end_time = ? WHERE user_username = ? AND status = 'in_progress'",
                (now, username)
            )
            
            # NO BORRAMOS NADA de stock_counts. Los datos están seguros.
            await conn.commit()
    
    return RedirectResponse(url='/admin/users', status_code=status.HTTP_302_FOUND)

@app.get('/admin/logout')
async def admin_logout(request: Request):
    response = RedirectResponse(url='/', status_code=status.HTTP_302_FOUND)
    response.delete_cookie("admin_logged_in")
    return response

# --- Lógica de inicialización ---
@app.on_event("startup")
async def startup_event():
    await init_db()
    await load_csv_data()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)